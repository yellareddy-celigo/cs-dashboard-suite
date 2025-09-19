#!/usr/bin/env python3
"""
Clear Data Dashboard Creator
Reads data from your Google Sheet and creates a clear, professional dashboard
Similar to your existing dashboard structure but in Excel format
"""

import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.drawing.image import Image
import json

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ClearDashboardCreator:
    def __init__(self):
        """Initialize the dashboard creator"""
        self.cs_data = None
        self.pre_data = None
        self.combined_data = None
        
    def connect_to_google_sheets(self):
        """Connect to your Google Sheet"""
        try:
            scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
            credentials = Credentials.from_service_account_file('service_account_key.json', scopes=scopes)
            self.gc = gspread.authorize(credentials)
            
            sheet_id = '1JN9HWj8JvClgTDe77x3KIxuyNUTsybPhYhhfzKy6Udc'
            self.sheet = self.gc.open_by_key(sheet_id)
            logger.info("✅ Connected to Google Sheets successfully")
            return True
        except Exception as e:
            logger.error(f"❌ Google Sheets connection failed: {e}")
            return False
    
    def load_data(self):
        """Load data from your Google Sheet"""
        try:
            # Load CS data
            cs_sheet = self.sheet.worksheet('Raw Data - CS')
            cs_values = cs_sheet.get_all_values()
            cs_headers = cs_values[0]
            cs_data = cs_values[1:]
            
            # Create DataFrame with unique headers
            cs_df_data = []
            for row in cs_data:
                if len(row) >= len(cs_headers):
                    cs_df_data.append(row[:len(cs_headers)])
                else:
                    # Pad with empty strings if row is shorter
                    padded_row = row + [''] * (len(cs_headers) - len(row))
                    cs_df_data.append(padded_row)
            
            self.cs_data = pd.DataFrame(cs_df_data, columns=cs_headers)
            logger.info(f"✅ Loaded {len(self.cs_data)} CS issues")
            
            # Load PRE data
            pre_sheet = self.sheet.worksheet('Raw Data - PRE')
            pre_values = pre_sheet.get_all_values()
            pre_headers = pre_values[0]
            pre_data = pre_values[1:]
            
            # Create DataFrame with unique headers
            pre_df_data = []
            for row in pre_data:
                if len(row) >= len(pre_headers):
                    pre_df_data.append(row[:len(pre_headers)])
                else:
                    # Pad with empty strings if row is shorter
                    padded_row = row + [''] * (len(pre_headers) - len(row))
                    pre_df_data.append(padded_row)
            
            self.pre_data = pd.DataFrame(pre_df_data, columns=pre_headers)
            logger.info(f"✅ Loaded {len(self.pre_data)} PRE issues")
            
            # Combine data
            self.cs_data['Project Type'] = 'CS'
            self.pre_data['Project Type'] = 'PRE'
            
            # Select key columns for analysis
            key_columns = ['Summary', 'Issue key', 'Issue Type', 'Status', 'Project Type', 'Created', 'Updated', 'Resolved', 'Priority', 'Assignee', 'Reporter']
            
            # Filter to available columns
            available_cs_cols = [col for col in key_columns if col in self.cs_data.columns]
            available_pre_cols = [col for col in key_columns if col in self.pre_data.columns]
            
            self.cs_analysis = self.cs_data[available_cs_cols].copy()
            self.pre_analysis = self.pre_data[available_pre_cols].copy()
            
            # Combine for overall analysis
            common_cols = list(set(available_cs_cols) & set(available_pre_cols))
            self.combined_data = pd.concat([
                self.cs_analysis[common_cols],
                self.pre_analysis[common_cols]
            ], ignore_index=True)
            
            logger.info(f"✅ Combined data: {len(self.combined_data)} total issues")
            return True
            
        except Exception as e:
            logger.error(f"❌ Error loading data: {e}")
            return False
    
    def process_dates(self, df):
        """Process date columns"""
        date_columns = ['Created', 'Updated', 'Resolved']
        
        for col in date_columns:
            if col in df.columns:
                df[col + '_dt'] = pd.to_datetime(df[col], errors='coerce')
                df[col + '_year'] = df[col + '_dt'].dt.year
                df[col + '_month'] = df[col + '_dt'].dt.month
                df[col + '_month_name'] = df[col + '_dt'].dt.strftime('%B')
                df[col + '_quarter'] = df[col + '_dt'].dt.quarter
        
        return df
    
    def generate_analysis(self):
        """Generate comprehensive analysis"""
        analysis = {}
        
        # Process dates
        self.cs_analysis = self.process_dates(self.cs_analysis)
        self.pre_analysis = self.process_dates(self.pre_analysis)
        self.combined_data = self.process_dates(self.combined_data)
        
        # Basic statistics
        analysis['total_issues'] = len(self.combined_data)
        analysis['cs_issues'] = len(self.cs_analysis)
        analysis['pre_issues'] = len(self.pre_analysis)
        
        # Status analysis
        analysis['status_distribution'] = self.combined_data['Status'].value_counts().to_dict()
        analysis['cs_status_distribution'] = self.cs_analysis['Status'].value_counts().to_dict()
        analysis['pre_status_distribution'] = self.pre_analysis['Status'].value_counts().to_dict()
        
        # Priority analysis
        if 'Priority' in self.combined_data.columns:
            analysis['priority_distribution'] = self.combined_data['Priority'].value_counts().to_dict()
        
        # Monthly trends
        if 'Created_month_name' in self.combined_data.columns:
            analysis['monthly_trends'] = self.combined_data['Created_month_name'].value_counts().to_dict()
            analysis['cs_monthly_trends'] = self.cs_analysis['Created_month_name'].value_counts().to_dict()
            analysis['pre_monthly_trends'] = self.pre_analysis['Created_month_name'].value_counts().to_dict()
        
        # Yearly trends
        if 'Created_year' in self.combined_data.columns:
            analysis['yearly_trends'] = self.combined_data['Created_year'].value_counts().to_dict()
        
        # Assignee analysis
        if 'Assignee' in self.combined_data.columns:
            analysis['top_assignees'] = self.combined_data['Assignee'].value_counts().head(10).to_dict()
        
        # Resolution analysis
        if 'Resolved' in self.combined_data.columns:
            resolved_data = self.combined_data[self.combined_data['Resolved'].notna()]
            analysis['resolved_issues'] = len(resolved_data)
            analysis['resolution_rate'] = (len(resolved_data) / len(self.combined_data)) * 100 if len(self.combined_data) > 0 else 0
        
        # Recent activity (last 30 days)
        if 'Created_dt' in self.combined_data.columns:
            thirty_days_ago = datetime.now() - timedelta(days=30)
            recent_issues = self.combined_data[self.combined_data['Created_dt'] > thirty_days_ago]
            analysis['recent_issues_30d'] = len(recent_issues)
        
        logger.info("✅ Analysis completed")
        return analysis
    
    def create_excel_dashboard(self, analysis, filename=None):
        """Create a clear Excel dashboard similar to your Google Sheet"""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'clear_dashboard_{timestamp}.xlsx'
        
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=14)
            subheader_font = Font(bold=True, color="FFFFFF", size=12)
            title_font = Font(bold=True, size=16, color="FFFFFF")
            
            header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
            subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            title_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            cs_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
            pre_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # 1. EXECUTIVE SUMMARY DASHBOARD
            ws_dashboard = wb.create_sheet("🎯 Executive Dashboard")
            
            dashboard_data = [
                ['🎯 CLEAR DATA DASHBOARD - EXECUTIVE SUMMARY', '', '', ''],
                ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '', ''],
                ['', '', '', ''],
                ['📊 KEY METRICS OVERVIEW', '', '', ''],
                ['Metric', 'Value', 'CS', 'PRE'],
                ['Total Issues', analysis['total_issues'], analysis['cs_issues'], analysis['pre_issues']],
                ['Resolved Issues', analysis.get('resolved_issues', 0), '', ''],
                ['Resolution Rate', f"{analysis.get('resolution_rate', 0):.1f}%", '', ''],
                ['Recent Issues (30d)', analysis.get('recent_issues_30d', 0), '', ''],
                ['', '', '', ''],
                ['📈 STATUS BREAKDOWN', '', '', ''],
                ['Status', 'Total', 'CS', 'PRE']
            ]
            
            # Add status breakdown
            for status, count in analysis['status_distribution'].items():
                cs_count = analysis['cs_status_distribution'].get(status, 0)
                pre_count = analysis['pre_status_distribution'].get(status, 0)
                dashboard_data.append([status, count, cs_count, pre_count])
            
            # Add monthly trends
            dashboard_data.extend([
                ['', '', '', ''],
                ['📅 MONTHLY TRENDS', '', '', ''],
                ['Month', 'Total', 'CS', 'PRE']
            ])
            
            if 'monthly_trends' in analysis:
                for month in ['January', 'February', 'March', 'April', 'May', 'June', 
                             'July', 'August', 'September', 'October', 'November', 'December']:
                    total = analysis['monthly_trends'].get(month, 0)
                    cs = analysis['cs_monthly_trends'].get(month, 0)
                    pre = analysis['pre_monthly_trends'].get(month, 0)
                    dashboard_data.append([month, total, cs, pre])
            
            # Write dashboard data
            for row in dashboard_data:
                ws_dashboard.append(row)
            
            # Style the dashboard
            for cell in ws_dashboard[1]:
                cell.font = title_font
                cell.fill = title_fill
                cell.border = border
            
            for cell in ws_dashboard[2]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.border = border
            
            # Style headers
            for row_num in [5, 11, 15]:  # Header rows
                for cell in ws_dashboard[row_num]:
                    cell.font = subheader_font
                    cell.fill = subheader_fill
                    cell.border = border
            
            # 2. DETAILED ANALYSIS SHEET
            ws_analysis = wb.create_sheet("📊 Detailed Analysis")
            
            analysis_data = [
                ['📊 DETAILED ISSUE ANALYSIS', '', '', ''],
                ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '', ''],
                ['', '', '', ''],
                ['🔍 CS vs PRE COMPARISON', '', '', ''],
                ['Metric', 'CS', 'PRE', 'Difference'],
                ['Total Issues', analysis['cs_issues'], analysis['pre_issues'], 
                 analysis['cs_issues'] - analysis['pre_issues']],
                ['', '', '', ''],
                ['📋 TOP ASSIGNEES', '', '', ''],
                ['Assignee', 'Total Issues', 'CS Issues', 'PRE Issues']
            ]
            
            if 'top_assignees' in analysis:
                for assignee, count in list(analysis['top_assignees'].items())[:10]:
                    cs_count = self.cs_analysis[self.cs_analysis['Assignee'] == assignee].shape[0]
                    pre_count = self.pre_analysis[self.pre_analysis['Assignee'] == assignee].shape[0]
                    analysis_data.append([assignee, count, cs_count, pre_count])
            
            # Add priority analysis
            if 'priority_distribution' in analysis:
                analysis_data.extend([
                    ['', '', '', ''],
                    ['⚡ PRIORITY DISTRIBUTION', '', '', ''],
                    ['Priority', 'Count', 'Percentage', '']
                ])
                
                for priority, count in analysis['priority_distribution'].items():
                    percentage = (count / analysis['total_issues']) * 100
                    analysis_data.append([priority, count, f"{percentage:.1f}%", ''])
            
            for row in analysis_data:
                ws_analysis.append(row)
            
            # Style the analysis sheet
            for cell in ws_analysis[1]:
                cell.font = title_font
                cell.fill = title_fill
                cell.border = border
            
            # 3. RAW DATA SHEET
            ws_raw = wb.create_sheet("📋 Raw Data")
            
            # Add CS data
            cs_headers = ['Project Type'] + list(self.cs_analysis.columns)
            cs_data_with_type = self.cs_analysis.copy()
            cs_data_with_type.insert(0, 'Project Type', 'CS')
            
            for r in dataframe_to_rows(cs_data_with_type[cs_headers], index=False, header=True):
                ws_raw.append(r)
            
            # Add PRE data
            pre_headers = ['Project Type'] + list(self.pre_analysis.columns)
            pre_data_with_type = self.pre_analysis.copy()
            pre_data_with_type.insert(0, 'Project Type', 'PRE')
            
            for r in dataframe_to_rows(pre_data_with_type[pre_headers], index=False, header=True):
                ws_raw.append(r)
            
            # Style raw data
            for cell in ws_raw[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Color code by project type
            for row_idx, row in enumerate(ws_raw.iter_rows(min_row=2), 2):
                project_type = ws_raw.cell(row=row_idx, column=1).value
                for cell in row:
                    if project_type == 'CS':
                        cell.fill = cs_fill
                    elif project_type == 'PRE':
                        cell.fill = pre_fill
                    cell.border = border
            
            # 4. VISUAL CHARTS SHEET
            ws_charts = wb.create_sheet("📈 Visual Charts")
            
            charts_data = [
                ['📈 VISUAL CHARTS & INSIGHTS', '', '', ''],
                ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '', ''],
                ['', '', '', ''],
                ['📊 MONTHLY ISSUE DISTRIBUTION', '', '', ''],
                ['Month', 'Total Issues', 'CS Issues', 'PRE Issues', 'Chart']
            ]
            
            if 'monthly_trends' in analysis:
                for month in ['January', 'February', 'March', 'April', 'May', 'June', 
                             'July', 'August', 'September', 'October', 'November', 'December']:
                    total = analysis['monthly_trends'].get(month, 0)
                    cs = analysis['cs_monthly_trends'].get(month, 0)
                    pre = analysis['pre_monthly_trends'].get(month, 0)
                    
                    # Create ASCII chart
                    max_val = max(analysis['monthly_trends'].values()) if analysis['monthly_trends'] else 1
                    bar_length = int((total / max_val) * 20) if max_val > 0 else 0
                    chart_bar = '🟨' * bar_length + '⬜' * (20 - bar_length)
                    
                    charts_data.append([month, total, cs, pre, chart_bar])
            
            charts_data.extend([
                ['', '', '', '', ''],
                ['📋 STATUS DISTRIBUTION', '', '', '', ''],
                ['Status', 'Count', 'Percentage', 'CS', 'PRE']
            ])
            
            for status, count in analysis['status_distribution'].items():
                percentage = (count / analysis['total_issues']) * 100
                cs_count = analysis['cs_status_distribution'].get(status, 0)
                pre_count = analysis['pre_status_distribution'].get(status, 0)
                charts_data.append([status, count, f"{percentage:.1f}%", cs_count, pre_count])
            
            for row in charts_data:
                ws_charts.append(row)
            
            # Style charts sheet
            for cell in ws_charts[1]:
                cell.font = title_font
                cell.fill = title_fill
                cell.border = border
            
            # Save the file
            wb.save(filename)
            logger.info(f"✅ Excel dashboard saved: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"❌ Error creating Excel dashboard: {e}")
            return None
    
    def run(self):
        """Run the complete dashboard creation process"""
        logger.info("🚀 Starting Clear Dashboard Creation")
        logger.info("=" * 50)
        
        # Connect to Google Sheets
        if not self.connect_to_google_sheets():
            return False
        
        # Load data
        if not self.load_data():
            return False
        
        # Generate analysis
        analysis = self.generate_analysis()
        
        # Create Excel dashboard
        filename = self.create_excel_dashboard(analysis)
        
        if filename:
            logger.info("✅ Clear dashboard created successfully!")
            logger.info(f"📁 File: {filename}")
            logger.info(f"📊 Analyzed {analysis['total_issues']} total issues")
            logger.info(f"   • CS Issues: {analysis['cs_issues']}")
            logger.info(f"   • PRE Issues: {analysis['pre_issues']}")
            logger.info(f"   • Resolution Rate: {analysis.get('resolution_rate', 0):.1f}%")
            return True
        else:
            logger.error("❌ Dashboard creation failed")
            return False

def main():
    """Main execution function"""
    creator = ClearDashboardCreator()
    success = creator.run()
    
    if success:
        print("\n🎉 SUCCESS! Clear dashboard created!")
        print("📋 Dashboard includes:")
        print("   • 🎯 Executive Dashboard - Key metrics and overview")
        print("   • 📊 Detailed Analysis - CS vs PRE comparison")
        print("   • 📋 Raw Data - All issues with color coding")
        print("   • 📈 Visual Charts - Monthly trends and status distribution")
        print("\n🔗 Open the Excel file to view your clear dashboard!")
    else:
        print("\n❌ Dashboard creation failed. Check the logs above for details.")
    
    return 0 if success else 1

if __name__ == "__main__":
    import sys
    sys.exit(main())
