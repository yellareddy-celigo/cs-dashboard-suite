#!/usr/bin/env python3
"""
JIRA MCP + Excel Analysis Tool
Uses JIRA MCP server to fetch data and export to Excel with holiday season analysis
Based on the Historical CS Analysis project requirements
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
import json

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class JiraMCPExcelAnalyzer:
    def __init__(self):
        """Initialize the analyzer"""
        self.data = None
        self.analysis = {}
        
        # Holiday season periods as defined in README
        self.holiday_periods = {
            'Black Friday Week': {'start': 'Nov 20', 'end': 'Nov 27'},
            'Cyber Monday': {'start': 'Nov 27', 'end': 'Dec 1'},
            'Holiday Shopping': {'start': 'Dec 1', 'end': 'Dec 24'},
            'Christmas Week': {'start': 'Dec 24', 'end': 'Jan 1'},
            'New Year Recovery': {'start': 'Jan 1', 'end': 'Jan 15'}
        }
    
    def classify_holiday_period(self, date_str):
        """Classify a date into holiday periods"""
        if not date_str:
            return 'Unknown'
        
        try:
            # Parse the date
            if 'T' in date_str:
                date_obj = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            else:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
            
            month = date_obj.month
            day = date_obj.day
            
            # Classify based on month and day
            if month == 11:  # November
                if day >= 20:
                    return 'Black Friday Week'
                elif day >= 27:
                    return 'Cyber Monday'
            elif month == 12:  # December
                if day <= 1:
                    return 'Cyber Monday'
                elif day <= 24:
                    return 'Holiday Shopping'
                else:
                    return 'Christmas Week'
            elif month == 1:  # January
                if day <= 15:
                    return 'New Year Recovery'
            
            return 'Off-Season'
            
        except Exception as e:
            logger.warning(f"Could not parse date: {date_str} - {e}")
            return 'Unknown'
    
    def process_jira_issues(self, issues):
        """Process JIRA issues from MCP into analysis format"""
        processed_data = []
        
        for issue in issues:
            fields = issue.get('fields', {})
            
            # Format dates
            def format_date(date_str):
                if date_str:
                    try:
                        dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                        return dt.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        return date_str
                return ''
            
            # Extract key information
            created_date = format_date(fields.get('created', ''))
            updated_date = format_date(fields.get('updated', ''))
            resolved_date = format_date(fields.get('resolutiondate', ''))
            
            row = {
                'JIRA ID': issue.get('key', ''),
                'Summary': fields.get('summary', ''),
                'Description': fields.get('description', ''),
                'Status': fields.get('status', {}).get('name', '') if fields.get('status') else '',
                'Priority': fields.get('priority', {}).get('name', '') if fields.get('priority') else '',
                'Issue Type': fields.get('issuetype', {}).get('name', '') if fields.get('issuetype') else '',
                'Created Date': created_date,
                'Updated Date': updated_date,
                'Resolved Date': resolved_date,
                'Assignee': fields.get('assignee', {}).get('displayName', 'Unassigned') if fields.get('assignee') else 'Unassigned',
                'Reporter': fields.get('reporter', {}).get('displayName', '') if fields.get('reporter') else '',
                'Project': fields.get('project', {}).get('name', '') if fields.get('project') else '',
                'Labels': ', '.join(fields.get('labels', [])) if fields.get('labels') else '',
                'Components': ', '.join([comp.get('name', '') for comp in fields.get('components', [])]) if fields.get('components') else '',
                'Resolution': fields.get('resolution', {}).get('name', '') if fields.get('resolution') else '',
                'Created Year': '',
                'Created Month': '',
                'Created Quarter': '',
                'Week Number': '',
                'Day of Week': '',
                'Holiday Period': '',
                'Resolution Time (Days)': '',
                'Is Holiday Season': False
            }
            
            # Add calculated fields
            if created_date:
                try:
                    created_dt = datetime.strptime(created_date, '%Y-%m-%d %H:%M:%S')
                    row['Created Year'] = created_dt.year
                    row['Created Month'] = created_dt.strftime('%B')
                    row['Created Quarter'] = f"Q{((created_dt.month - 1) // 3) + 1}"
                    row['Week Number'] = created_dt.isocalendar()[1]
                    row['Day of Week'] = created_dt.strftime('%A')
                    
                    # Holiday period classification
                    holiday_period = self.classify_holiday_period(created_date)
                    row['Holiday Period'] = holiday_period
                    row['Is Holiday Season'] = holiday_period != 'Off-Season' and holiday_period != 'Unknown'
                    
                    # Calculate resolution time
                    if resolved_date:
                        resolved_dt = datetime.strptime(resolved_date, '%Y-%m-%d %H:%M:%S')
                        resolution_time = (resolved_dt - created_dt).days
                        row['Resolution Time (Days)'] = resolution_time
                except Exception as e:
                    logger.warning(f"Error processing date for {issue.get('key', '')}: {e}")
            
            processed_data.append(row)
        
        df = pd.DataFrame(processed_data)
        logger.info(f"✅ Processed {len(df)} issues into DataFrame")
        return df
    
    def generate_holiday_analysis(self, df):
        """Generate holiday season analysis as per README requirements"""
        analysis = {}
        
        # Basic statistics
        analysis['total_issues'] = len(df)
        analysis['open_issues'] = len(df[~df['Status'].str.contains('Closed|Resolved|Done', case=False, na=False)])
        analysis['closed_issues'] = len(df[df['Status'].str.contains('Closed|Resolved|Done', case=False, na=False)])
        analysis['resolution_rate'] = (analysis['closed_issues'] / analysis['total_issues'] * 100) if analysis['total_issues'] > 0 else 0
        
        # Holiday season analysis
        holiday_df = df[df['Is Holiday Season'] == True]
        analysis['holiday_season_issues'] = len(holiday_df)
        analysis['holiday_season_percentage'] = (len(holiday_df) / len(df) * 100) if len(df) > 0 else 0
        
        # Holiday period breakdown
        analysis['holiday_period_distribution'] = df[df['Is Holiday Season'] == True]['Holiday Period'].value_counts().to_dict()
        
        # Status distribution
        analysis['status_distribution'] = df['Status'].value_counts().to_dict()
        
        # Priority distribution
        analysis['priority_distribution'] = df['Priority'].value_counts().to_dict()
        
        # Time-based analysis
        if 'Created Year' in df.columns and df['Created Year'].notna().any():
            analysis['yearly_trend'] = df.groupby('Created Year').size().to_dict()
        
        if 'Created Month' in df.columns and df['Created Month'].notna().any():
            analysis['monthly_pattern'] = df['Created Month'].value_counts().to_dict()
        
        # Holiday season vs off-season comparison
        analysis['holiday_vs_offseason'] = {
            'Holiday Season': len(holiday_df),
            'Off Season': len(df[df['Is Holiday Season'] == False])
        }
        
        # Resolution time analysis
        if 'Resolution Time (Days)' in df.columns:
            resolved_df = df[df['Resolution Time (Days)'].notna() & (df['Resolution Time (Days)'] != '')]
            if len(resolved_df) > 0:
                analysis['avg_resolution_time'] = resolved_df['Resolution Time (Days)'].mean()
                analysis['median_resolution_time'] = resolved_df['Resolution Time (Days)'].median()
                
                # Holiday season resolution times
                holiday_resolved = resolved_df[resolved_df['Is Holiday Season'] == True]
                if len(holiday_resolved) > 0:
                    analysis['holiday_avg_resolution_time'] = holiday_resolved['Resolution Time (Days)'].mean()
                    analysis['holiday_median_resolution_time'] = holiday_resolved['Resolution Time (Days)'].median()
        
        # Top assignees
        analysis['top_assignees'] = df['Assignee'].value_counts().head(10).to_dict()
        
        # Recent activity (last 30 days)
        try:
            df['Created_dt'] = pd.to_datetime(df['Created Date'], errors='coerce')
            thirty_days_ago = datetime.now() - timedelta(days=30)
            recent_cases = df[df['Created_dt'] > thirty_days_ago]
            analysis['recent_cases_30d'] = len(recent_cases)
        except:
            analysis['recent_cases_30d'] = 0
        
        logger.info("✅ Holiday season analysis completed")
        return analysis
    
    def export_to_excel(self, df, analysis, filename=None):
        """Export data and analysis to Excel file with holiday season focus"""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'jira_holiday_analysis_{timestamp}.xlsx'
        
        try:
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            holiday_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # 1. Raw Data Sheet
            ws_raw = wb.create_sheet("Raw Data")
            for r in dataframe_to_rows(df, index=False, header=True):
                ws_raw.append(r)
            
            # Style the header
            for cell in ws_raw[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Highlight holiday season rows
            for row_idx, row in enumerate(ws_raw.iter_rows(min_row=2), 2):
                if df.iloc[row_idx-2]['Is Holiday Season']:
                    for cell in row:
                        cell.fill = holiday_fill
            
            # 2. Holiday Season Analysis Sheet
            ws_holiday = wb.create_sheet("Holiday Season Analysis")
            holiday_data = [
                ['Holiday Season Analysis Dashboard', '', ''],
                ['Last Updated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ''],
                ['', '', ''],
                ['Key Metrics', '', ''],
                ['Total Issues:', analysis['total_issues'], ''],
                ['Holiday Season Issues:', analysis['holiday_season_issues'], f"{analysis['holiday_season_percentage']:.1f}%"],
                ['Open Issues:', analysis['open_issues'], f"{(analysis['open_issues']/analysis['total_issues']*100):.1f}%"],
                ['Closed Issues:', analysis['closed_issues'], f"{(analysis['closed_issues']/analysis['total_issues']*100):.1f}%"],
                ['Resolution Rate:', f"{analysis['resolution_rate']:.1f}%", ''],
                ['', '', ''],
                ['Holiday Period Breakdown', '', '']
            ]
            
            for period, count in analysis['holiday_period_distribution'].items():
                holiday_data.append([period, count, f"{(count/analysis['holiday_season_issues']*100):.1f}%"])
            
            holiday_data.extend([
                ['', '', ''],
                ['Holiday vs Off-Season', '', ''],
                ['Holiday Season:', analysis['holiday_vs_offseason']['Holiday Season'], ''],
                ['Off Season:', analysis['holiday_vs_offseason']['Off Season'], '']
            ])
            
            if 'holiday_avg_resolution_time' in analysis:
                holiday_data.extend([
                    ['', '', ''],
                    ['Resolution Times', '', ''],
                    ['Holiday Season Avg:', f"{analysis['holiday_avg_resolution_time']:.1f} days", ''],
                    ['Overall Avg:', f"{analysis['avg_resolution_time']:.1f} days", '']
                ])
            
            for row in holiday_data:
                ws_holiday.append(row)
            
            # Style the holiday analysis sheet
            for cell in ws_holiday[1]:
                cell.font = Font(bold=True, size=14, color="FFFFFF")
                cell.fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
            
            for cell in ws_holiday[2]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            
            # 3. General Analysis Sheet
            ws_analysis = wb.create_sheet("General Analysis")
            analysis_data = [
                ['General Analysis', '', ''],
                ['Last Updated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ''],
                ['', '', ''],
                ['Status Distribution', '', '']
            ]
            
            for status, count in analysis['status_distribution'].items():
                analysis_data.append([status, count, f"{(count/analysis['total_issues']*100):.1f}%"])
            
            analysis_data.extend([
                ['', '', ''],
                ['Priority Distribution', '', '']
            ])
            
            for priority, count in analysis['priority_distribution'].items():
                analysis_data.append([priority, count, f"{(count/analysis['total_issues']*100):.1f}%"])
            
            analysis_data.extend([
                ['', '', ''],
                ['Top Assignees', '', '']
            ])
            
            for assignee, count in analysis['top_assignees'].items():
                analysis_data.append([assignee, count, f"{(count/analysis['total_issues']*100):.1f}%"])
            
            for row in analysis_data:
                ws_analysis.append(row)
            
            # Style the analysis sheet
            for cell in ws_analysis[1]:
                cell.font = Font(bold=True, size=14, color="FFFFFF")
                cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            
            for cell in ws_analysis[2]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            
            # 4. Executive Dashboard Sheet
            ws_dashboard = wb.create_sheet("Executive Dashboard")
            dashboard_data = [
                ['JIRA Holiday Season Analysis - Executive Dashboard', '', ''],
                ['Last Updated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ''],
                ['', '', ''],
                ['Key Performance Indicators', '', ''],
                ['Total Issues Analyzed:', analysis['total_issues'], ''],
                ['Holiday Season Issues:', analysis['holiday_season_issues'], f"{analysis['holiday_season_percentage']:.1f}%"],
                ['Overall Resolution Rate:', f"{analysis['resolution_rate']:.1f}%", ''],
                ['Recent Activity (30 days):', analysis['recent_cases_30d'], ''],
                ['', '', ''],
                ['Holiday Season Impact', '', ''],
                ['Black Friday Week:', analysis['holiday_period_distribution'].get('Black Friday Week', 0), ''],
                ['Christmas Week:', analysis['holiday_period_distribution'].get('Christmas Week', 0), ''],
                ['Holiday Shopping:', analysis['holiday_period_distribution'].get('Holiday Shopping', 0), ''],
                ['', '', ''],
                ['View Details:', 'See Raw Data sheet', ''],
                ['View Holiday Analysis:', 'See Holiday Season Analysis sheet', ''],
                ['View General Analysis:', 'See General Analysis sheet', '']
            ]
            
            for row in dashboard_data:
                ws_dashboard.append(row)
            
            # Style the dashboard
            for cell in ws_dashboard[1]:
                cell.font = Font(bold=True, size=16, color="FFFFFF")
                cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            
            for cell in ws_dashboard[2]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            
            # Save the file
            wb.save(filename)
            logger.info(f"✅ Excel file saved: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"❌ Error saving Excel file: {e}")
            return None

def main():
    """Main execution function - this will be called by the MCP integration"""
    print("🚀 JIRA MCP + Excel Analyzer")
    print("=" * 50)
    print("This script is designed to work with JIRA MCP server")
    print("It will be called by the MCP integration to process JIRA data")
    print("and export to Excel with holiday season analysis.")
    print()
    print("To use this script:")
    print("1. Call it from the MCP integration")
    print("2. Pass JIRA issues data to the process_jira_issues() method")
    print("3. The script will generate Excel analysis with holiday season focus")
    
    return True

if __name__ == "__main__":
    main()
