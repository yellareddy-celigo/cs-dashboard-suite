#!/usr/bin/env python3
"""
Simple Clear Dashboard Creator
Creates a clean, professional dashboard from your Google Sheet data
"""

import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class SimpleDashboardCreator:
    def __init__(self):
        """Initialize the dashboard creator"""
        self.cs_data = None
        self.pre_data = None
        
    def connect_to_google_sheets(self):
        """Connect to your Google Sheet"""
        try:
            scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
            credentials = Credentials.from_service_account_file('service_account_key.json', scopes=scopes)
            self.gc = gspread.authorize(credentials)
            
            sheet_id = '1JN9HWj8JvClgTDe77x3KIxuyNUTsybPhYhhfzKy6Udc'
            self.sheet = self.gc.open_by_key(sheet_id)
            logger.info("✅ Connected to Google Sheets successfully")
            return True
        except Exception as e:
            logger.error(f"❌ Google Sheets connection failed: {e}")
            return False
    
    def load_data(self):
        """Load data from your Google Sheet"""
        try:
            # Load CS data
            cs_sheet = self.sheet.worksheet('Raw Data - CS')
            cs_values = cs_sheet.get_all_values()
            cs_headers = cs_values[0]
            cs_data = cs_values[1:]
            
            # Create DataFrame
            self.cs_data = pd.DataFrame(cs_data, columns=cs_headers)
            logger.info(f"✅ Loaded {len(self.cs_data)} CS issues")
            
            # Load PRE data
            pre_sheet = self.sheet.worksheet('Raw Data - PRE')
            pre_values = pre_sheet.get_all_values()
            pre_headers = pre_values[0]
            pre_data = pre_values[1:]
            
            # Create DataFrame
            self.pre_data = pd.DataFrame(pre_data, columns=pre_headers)
            logger.info(f"✅ Loaded {len(self.pre_data)} PRE issues")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Error loading data: {e}")
            return False
    
    def process_dates(self, df):
        """Process date columns"""
        date_columns = ['Created', 'Updated', 'Resolved']
        
        for col in date_columns:
            if col in df.columns:
                try:
                    df[col + '_dt'] = pd.to_datetime(df[col], errors='coerce')
                    df[col + '_year'] = df[col + '_dt'].dt.year
                    df[col + '_month'] = df[col + '_dt'].dt.month
                    df[col + '_month_name'] = df[col + '_dt'].dt.strftime('%B')
                except:
                    pass
        
        return df
    
    def generate_analysis(self):
        """Generate comprehensive analysis"""
        analysis = {}
        
        # Process dates
        self.cs_data = self.process_dates(self.cs_data)
        self.pre_data = self.process_dates(self.pre_data)
        
        # Basic statistics
        analysis['total_issues'] = len(self.cs_data) + len(self.pre_data)
        analysis['cs_issues'] = len(self.cs_data)
        analysis['pre_issues'] = len(self.pre_data)
        
        # Status analysis
        cs_status = self.cs_data['Status'].value_counts().to_dict() if 'Status' in self.cs_data.columns else {}
        pre_status = self.pre_data['Status'].value_counts().to_dict() if 'Status' in self.pre_data.columns else {}
        
        # Combine status counts
        all_status = {}
        for status, count in cs_status.items():
            all_status[status] = all_status.get(status, 0) + count
        for status, count in pre_status.items():
            all_status[status] = all_status.get(status, 0) + count
        
        analysis['status_distribution'] = all_status
        analysis['cs_status_distribution'] = cs_status
        analysis['pre_status_distribution'] = pre_status
        
        # Monthly trends
        cs_monthly = self.cs_data['Created_month_name'].value_counts().to_dict() if 'Created_month_name' in self.cs_data.columns else {}
        pre_monthly = self.pre_data['Created_month_name'].value_counts().to_dict() if 'Created_month_name' in self.pre_data.columns else {}
        
        # Combine monthly counts
        all_monthly = {}
        for month, count in cs_monthly.items():
            all_monthly[month] = all_monthly.get(month, 0) + count
        for month, count in pre_monthly.items():
            all_monthly[month] = all_monthly.get(month, 0) + count
        
        analysis['monthly_trends'] = all_monthly
        analysis['cs_monthly_trends'] = cs_monthly
        analysis['pre_monthly_trends'] = pre_monthly
        
        # Priority analysis
        if 'Priority' in self.cs_data.columns:
            cs_priority = self.cs_data['Priority'].value_counts().to_dict()
            analysis['cs_priority_distribution'] = cs_priority
        
        if 'Priority' in self.pre_data.columns:
            pre_priority = self.pre_data['Priority'].value_counts().to_dict()
            analysis['pre_priority_distribution'] = pre_priority
        
        # Assignee analysis
        if 'Assignee' in self.cs_data.columns:
            cs_assignee = self.cs_data['Assignee'].value_counts().head(10).to_dict()
            analysis['cs_top_assignees'] = cs_assignee
        
        if 'Assignee' in self.pre_data.columns:
            pre_assignee = self.pre_data['Assignee'].value_counts().head(10).to_dict()
            analysis['pre_top_assignees'] = pre_assignee
        
        logger.info("✅ Analysis completed")
        return analysis
    
    def create_excel_dashboard(self, analysis, filename=None):
        """Create a clear Excel dashboard"""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'clear_dashboard_{timestamp}.xlsx'
        
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Define styles
            title_font = Font(bold=True, size=16, color="FFFFFF")
            header_font = Font(bold=True, size=12, color="FFFFFF")
            subheader_font = Font(bold=True, size=11, color="FFFFFF")
            
            title_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
            subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cs_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
            pre_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # 1. EXECUTIVE DASHBOARD
            ws_dashboard = wb.create_sheet("🎯 Executive Dashboard")
            
            dashboard_data = [
                ['🎯 CLEAR DATA DASHBOARD - EXECUTIVE SUMMARY', '', '', ''],
                ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '', ''],
                ['', '', '', ''],
                ['📊 KEY METRICS OVERVIEW', '', '', ''],
                ['Metric', 'Value', 'CS', 'PRE'],
                ['Total Issues', analysis['total_issues'], analysis['cs_issues'], analysis['pre_issues']],
                ['', '', '', ''],
                ['📈 STATUS BREAKDOWN', '', '', ''],
                ['Status', 'Total', 'CS', 'PRE']
            ]
            
            # Add status breakdown
            for status, count in analysis['status_distribution'].items():
                cs_count = analysis['cs_status_distribution'].get(status, 0)
                pre_count = analysis['pre_status_distribution'].get(status, 0)
                dashboard_data.append([status, count, cs_count, pre_count])
            
            # Add monthly trends
            dashboard_data.extend([
                ['', '', '', ''],
                ['📅 MONTHLY TRENDS', '', '', ''],
                ['Month', 'Total', 'CS', 'PRE']
            ])
            
            months = ['January', 'February', 'March', 'April', 'May', 'June', 
                     'July', 'August', 'September', 'October', 'November', 'December']
            
            for month in months:
                total = analysis['monthly_trends'].get(month, 0)
                cs = analysis['cs_monthly_trends'].get(month, 0)
                pre = analysis['pre_monthly_trends'].get(month, 0)
                dashboard_data.append([month, total, cs, pre])
            
            # Write dashboard data
            for row in dashboard_data:
                ws_dashboard.append(row)
            
            # Style the dashboard
            for cell in ws_dashboard[1]:
                cell.font = title_font
                cell.fill = title_fill
                cell.border = border
            
            # Style headers
            for row_num in [5, 9, 15]:  # Header rows
                for cell in ws_dashboard[row_num]:
                    cell.font = subheader_font
                    cell.fill = subheader_fill
                    cell.border = border
            
            # 2. CS DETAILED ANALYSIS
            ws_cs = wb.create_sheet("📊 CS Analysis")
            
            cs_data = [
                ['📊 CUSTOMER SUCCESS (CS) DETAILED ANALYSIS', '', '', ''],
                ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '', ''],
                ['Total CS Issues:', analysis['cs_issues'], '', ''],
                ['', '', '', ''],
                ['📋 CS STATUS BREAKDOWN', '', '', ''],
                ['Status', 'Count', 'Percentage', '']
            ]
            
            for status, count in analysis['cs_status_distribution'].items():
                percentage = (count / analysis['cs_issues']) * 100 if analysis['cs_issues'] > 0 else 0
                cs_data.append([status, count, f"{percentage:.1f}%", ''])
            
            # Add monthly trends
            cs_data.extend([
                ['', '', '', ''],
                ['📅 CS MONTHLY TRENDS', '', '', ''],
                ['Month', 'Count', 'Percentage', '']
            ])
            
            for month in months:
                count = analysis['cs_monthly_trends'].get(month, 0)
                percentage = (count / analysis['cs_issues']) * 100 if analysis['cs_issues'] > 0 else 0
                cs_data.append([month, count, f"{percentage:.1f}%", ''])
            
            # Add priority analysis
            if 'cs_priority_distribution' in analysis:
                cs_data.extend([
                    ['', '', '', ''],
                    ['⚡ CS PRIORITY BREAKDOWN', '', '', ''],
                    ['Priority', 'Count', 'Percentage', '']
                ])
                
                for priority, count in analysis['cs_priority_distribution'].items():
                    percentage = (count / analysis['cs_issues']) * 100 if analysis['cs_issues'] > 0 else 0
                    cs_data.append([priority, count, f"{percentage:.1f}%", ''])
            
            for row in cs_data:
                ws_cs.append(row)
            
            # Style CS sheet
            for cell in ws_cs[1]:
                cell.font = title_font
                cell.fill = title_fill
                cell.border = border
            
            # 3. PRE DETAILED ANALYSIS
            ws_pre = wb.create_sheet("🔧 PRE Analysis")
            
            pre_data = [
                ['🔧 PRE (PREBUILT) DETAILED ANALYSIS', '', '', ''],
                ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '', ''],
                ['Total PRE Issues:', analysis['pre_issues'], '', ''],
                ['', '', '', ''],
                ['📋 PRE STATUS BREAKDOWN', '', '', ''],
                ['Status', 'Count', 'Percentage', '']
            ]
            
            for status, count in analysis['pre_status_distribution'].items():
                percentage = (count / analysis['pre_issues']) * 100 if analysis['pre_issues'] > 0 else 0
                pre_data.append([status, count, f"{percentage:.1f}%", ''])
            
            # Add monthly trends
            pre_data.extend([
                ['', '', '', ''],
                ['📅 PRE MONTHLY TRENDS', '', '', ''],
                ['Month', 'Count', 'Percentage', '']
            ])
            
            for month in months:
                count = analysis['pre_monthly_trends'].get(month, 0)
                percentage = (count / analysis['pre_issues']) * 100 if analysis['pre_issues'] > 0 else 0
                pre_data.append([month, count, f"{percentage:.1f}%", ''])
            
            # Add priority analysis
            if 'pre_priority_distribution' in analysis:
                pre_data.extend([
                    ['', '', '', ''],
                    ['⚡ PRE PRIORITY BREAKDOWN', '', '', ''],
                    ['Priority', 'Count', 'Percentage', '']
                ])
                
                for priority, count in analysis['pre_priority_distribution'].items():
                    percentage = (count / analysis['pre_issues']) * 100 if analysis['pre_issues'] > 0 else 0
                    pre_data.append([priority, count, f"{percentage:.1f}%", ''])
            
            for row in pre_data:
                ws_pre.append(row)
            
            # Style PRE sheet
            for cell in ws_pre[1]:
                cell.font = title_font
                cell.fill = title_fill
                cell.border = border
            
            # 4. RAW DATA SHEET
            ws_raw = wb.create_sheet("📋 Raw Data")
            
            # Add CS data
            cs_headers = ['Project Type'] + list(self.cs_data.columns)
            cs_data_with_type = self.cs_data.copy()
            cs_data_with_type['Project Type'] = 'CS'
            
            # Select key columns for display
            key_columns = ['Project Type', 'Summary', 'Issue key', 'Status', 'Priority', 'Created', 'Updated', 'Assignee']
            available_columns = [col for col in key_columns if col in cs_data_with_type.columns]
            
            for r in dataframe_to_rows(cs_data_with_type[available_columns], index=False, header=True):
                ws_raw.append(r)
            
            # Add PRE data
            pre_data_with_type = self.pre_data.copy()
            pre_data_with_type['Project Type'] = 'PRE'
            
            available_pre_columns = [col for col in key_columns if col in pre_data_with_type.columns]
            
            for r in dataframe_to_rows(pre_data_with_type[available_pre_columns], index=False, header=False):
                ws_raw.append(r)
            
            # Style raw data
            for cell in ws_raw[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Color code by project type
            for row_idx, row in enumerate(ws_raw.iter_rows(min_row=2), 2):
                project_type = ws_raw.cell(row=row_idx, column=1).value
                for cell in row:
                    if project_type == 'CS':
                        cell.fill = cs_fill
                    elif project_type == 'PRE':
                        cell.fill = pre_fill
                    cell.border = border
            
            # Save the file
            wb.save(filename)
            logger.info(f"✅ Excel dashboard saved: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"❌ Error creating Excel dashboard: {e}")
            return None
    
    def run(self):
        """Run the complete dashboard creation process"""
        logger.info("🚀 Starting Simple Clear Dashboard Creation")
        logger.info("=" * 50)
        
        # Connect to Google Sheets
        if not self.connect_to_google_sheets():
            return False
        
        # Load data
        if not self.load_data():
            return False
        
        # Generate analysis
        analysis = self.generate_analysis()
        
        # Create Excel dashboard
        filename = self.create_excel_dashboard(analysis)
        
        if filename:
            logger.info("✅ Clear dashboard created successfully!")
            logger.info(f"📁 File: {filename}")
            logger.info(f"📊 Analyzed {analysis['total_issues']} total issues")
            logger.info(f"   • CS Issues: {analysis['cs_issues']}")
            logger.info(f"   • PRE Issues: {analysis['pre_issues']}")
            return True
        else:
            logger.error("❌ Dashboard creation failed")
            return False

def main():
    """Main execution function"""
    creator = SimpleDashboardCreator()
    success = creator.run()
    
    if success:
        print("\n🎉 SUCCESS! Clear dashboard created!")
        print("📋 Dashboard includes:")
        print("   • 🎯 Executive Dashboard - Key metrics and overview")
        print("   • 📊 CS Analysis - Customer Success detailed analysis")
        print("   • 🔧 PRE Analysis - Prebuilt detailed analysis")
        print("   • 📋 Raw Data - All issues with color coding")
        print("\n🔗 Open the Excel file to view your clear dashboard!")
    else:
        print("\n❌ Dashboard creation failed. Check the logs above for details.")
    
    return 0 if success else 1

if __name__ == "__main__":
    import sys
    sys.exit(main())
