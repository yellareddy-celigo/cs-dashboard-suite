#!/usr/bin/env python3
"""
DEEP DIVE DETAILED ANALYSIS
Extracts specific flows, mappings, fields, and error details from case descriptions and comments
Provides actionable, detailed information for each integration app

Usage:
    python3 deep_dive_detailed_analysis.py --file 2025.csv --output Deep_Dive_Details.xlsx
"""

import pandas as pd
import argparse
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter

VERSION = "1.0.0"

# ============================================================================
# EXTRACTION FUNCTIONS
# ============================================================================

def extract_flow_names(text):
    """Extract flow names from text - improved to get cleaner, more specific flow names."""
    if pd.isna(text):
        return []
    
    text_str = str(text)
    flows = []
    
    # Look for explicit flow names in brackets or quotes
    explicit_patterns = [
        r'flow\s*name[:\s]+([^\n\]]+?)[\]\n]',  # Flow Name: ...
        r'\[([^\]]{10,80})\]',  # Text in brackets (often flow names)
        r'"([^"]{10,80}(?:flow|import|export|sync)[^"]{0,20})"',  # Quoted flow names
    ]
    
    for pattern in explicit_patterns:
        matches = re.findall(pattern, text_str, re.IGNORECASE)
        flows.extend(matches)
    
    # Common flow patterns with context
    flow_patterns = [
        r'(?:in|from|at)\s+(?:the\s+)?([a-zA-Z0-9\s\-]+\s+(?:to|from)\s+[a-zA-Z0-9\s\-]+)\s+flow',
        r'([a-zA-Z]+\s+(?:to|from)\s+[a-zA-Z]+\s+(?:flow|import|export|sync))',
        r'(?:the|a)\s+([a-zA-Z0-9\s]+(?:import|export|sync|flow))\s+(?:is|has|was|flow)',
    ]
    
    for pattern in flow_patterns:
        matches = re.findall(pattern, text_str, re.IGNORECASE)
        flows.extend(matches)
    
    # Specific flow type keywords (high confidence)
    flow_keywords = {
        'order': ['order to netsuite', 'netsuite to order', 'order import', 'order export', 
                  'sales order', 'purchase order', 'order sync'],
        'product': ['product import', 'product export', 'product sync', 'item sync'],
        'inventory': ['inventory sync', 'inventory import', 'inventory export'],
        'customer': ['customer sync', 'customer import', 'customer export'],
        'fulfillment': ['item fulfillment', 'fulfillment sync', 'fulfillment import'],
        'settlement': ['settlement', 'settlement import', 'settlement report'],
        'shipment': ['shipment', 'shipment import', 'shipment export'],
        'refund': ['refund', 'refund import', 'refund sync'],
        'payment': ['payment sync', 'payment import', 'customer payment'],
        'invoice': ['invoice sync', 'invoice import', 'invoice export'],
        'cash sale': ['cash sale', 'cash sale import'],
        'credit memo': ['credit memo', 'credit memo import'],
    }
    
    text_lower = text_str.lower()
    for category, keywords in flow_keywords.items():
        for keyword in keywords:
            if keyword in text_lower:
                flows.append(keyword)
    
    # Clean and deduplicate
    cleaned_flows = []
    seen_flows = set()
    
    for flow in flows:
        flow = flow.strip().strip(':').strip(',').strip(']').strip('[')
        flow = re.sub(r'\s+', ' ', flow)  # Normalize whitespace
        
        # Skip if too short or too generic
        if len(flow) < 5:
            continue
        
        # Skip generic words
        generic = ['the flow', 'a flow', 'this flow', 'run the', 'failed to', 
                   'unable to', 'trying to', 'want to', 'need to']
        if any(g in flow.lower() for g in generic):
            continue
        
        # Check for duplicates (case-insensitive)
        flow_lower = flow.lower()
        if flow_lower not in seen_flows:
            cleaned_flows.append(flow)
            seen_flows.add(flow_lower)
    
    return cleaned_flows[:8]  # Limit to 8 most relevant

def extract_field_mappings(text):
    """Extract field and mapping issues from text."""
    if pd.isna(text):
        return []
    
    text_str = str(text)
    mappings = []
    
    # Field mapping patterns
    field_patterns = [
        r'field[:\s]+([a-zA-Z0-9_\s]+?)(?=\s+(?:is|not|missing|error|fail|$))',
        r'mapping[:\s]+([^,.\n]+?)(?=\s+(?:is|not|missing|error|fail|$))',
        r'(?:missing|undefined|null)\s+(?:field|value|mapping)[:\s]+([a-zA-Z0-9_\s]+)',
        r'([a-zA-Z0-9_]+)\s+field\s+(?:is|not|missing|error)',
        r'custom\s+field[:\s]+([a-zA-Z0-9_\s]+)',
        r'(?:netsuite|shopify|salesforce|amazon)\s+field[:\s]+([a-zA-Z0-9_\s]+)',
    ]
    
    for pattern in field_patterns:
        matches = re.findall(pattern, text_str, re.IGNORECASE)
        mappings.extend(matches)
    
    # Clean and deduplicate
    cleaned_mappings = []
    for mapping in mappings:
        mapping = mapping.strip().strip(':').strip(',')
        if len(mapping) > 2 and mapping not in cleaned_mappings:
            cleaned_mappings.append(mapping)
    
    return cleaned_mappings[:15]  # Limit to 15 most relevant

def extract_error_messages(text):
    """Extract specific error messages from text - improved to get clean, complete errors."""
    if pd.isna(text):
        return []
    
    text_str = str(text)
    errors = []
    
    # Look for complete error messages with better patterns
    error_patterns = [
        # Error with full message in quotes
        r'[Ee]rror[:\s]+"([^"]{20,200})"',
        # Error: message pattern (capture until period or newline)
        r'[Ee]rror[:\s]+([A-Z][^.\n]{20,200}[.!])',
        # Exception patterns
        r'[Ee]xception[:\s]+([A-Z][^.\n]{20,200}[.!])',
        # Status code errors
        r'([Ss]tatus [Cc]ode[:\s]+\d{3}[^.\n]{0,100})',
        # Failed to... patterns
        r'([Ff]ailed to [^.\n]{10,150}[.!])',
        # Unable to... patterns
        r'([Uu]unable to [^.\n]{10,150}[.!])',
        # Cannot... patterns
        r'([Cc]annot [^.\n]{10,150}[.!])',
        # Specific error formats
        r'((?:Invalid|Missing|Undefined)[^.\n]{10,150}[.!])',
        # Hook/function errors
        r'(hook (?:function )?error[^.\n]{10,150})',
        # Integration-specific errors
        r'(Integration (?:is )?corrupted[^.\n]{0,100})',
    ]
    
    for pattern in error_patterns:
        matches = re.findall(pattern, text_str)
        errors.extend(matches)
    
    # Clean and deduplicate
    cleaned_errors = []
    seen_errors = set()
    
    for error in errors:
        error = error.strip().strip(':').strip(',').strip('"')
        
        # Skip if too short, too long, or generic
        if len(error) < 20 or len(error) > 200:
            continue
        
        # Skip generic phrases
        generic_phrases = ['still persists', 'not working', 'issue', 'problem', 
                          'please', 'thank you', 'steps to reproduce']
        if any(phrase in error.lower() for phrase in generic_phrases):
            continue
        
        # Skip if it's just a fragment (no verb or too few words)
        word_count = len(error.split())
        if word_count < 4:
            continue
        
        # Check for similarity with existing errors (avoid duplicates)
        error_lower = error.lower()
        is_duplicate = False
        for seen in seen_errors:
            if error_lower in seen or seen in error_lower:
                is_duplicate = True
                break
        
        if not is_duplicate:
            cleaned_errors.append(error)
            seen_errors.add(error_lower)
    
    return cleaned_errors[:5]  # Limit to 5 most relevant, high-quality errors

def extract_pre_prd_references(text):
    """Extract PRE/PRD references from text."""
    if pd.isna(text):
        return []
    
    text_str = str(text)
    
    # PRE/PRD patterns
    pre_prd_pattern = r'(PR[ED]-\d+)'
    matches = re.findall(pre_prd_pattern, text_str, re.IGNORECASE)
    
    return list(set(matches))[:20]  # Deduplicate and limit

def extract_record_types(text):
    """Extract NetSuite/system record types mentioned."""
    if pd.isna(text):
        return []
    
    text_str = str(text).lower()
    
    record_types = [
        'sales order', 'purchase order', 'customer', 'item', 'invoice',
        'cash sale', 'item fulfillment', 'item receipt', 'vendor bill',
        'credit memo', 'customer deposit', 'journal entry', 'inventory adjustment',
        'transfer order', 'assembly build', 'work order', 'opportunity',
        'estimate', 'return authorization', 'vendor payment', 'customer payment'
    ]
    
    found_types = []
    for record_type in record_types:
        if record_type in text_str and record_type not in found_types:
            found_types.append(record_type)
    
    return found_types[:10]

# ============================================================================
# DEEP DIVE ANALYSIS
# ============================================================================

def deep_dive_analysis(csv_file, output_file=None):
    """Perform detailed deep dive analysis."""
    
    print("="*100)
    print(f"DEEP DIVE DETAILED ANALYSIS")
    print(f"Version: {VERSION}")
    print("="*100)
    
    # Load CSV
    df = pd.read_csv(csv_file)
    print(f"\n✅ Loaded {len(df)} cases")
    
    # Identify columns
    key_col = 'Issue key'
    summary_col = 'Summary'
    description_col = 'Description'
    case_type_col = [col for col in df.columns if 'case type' in col.lower() and 'custom field' in col.lower()][0]
    integration_col = [col for col in df.columns if 'integration app' in col.lower() and 'custom field' in col.lower()][0]
    priority_col = 'Priority'
    status_col = 'Status'
    resolution_col = 'Resolution'
    
    # Get all comment columns
    comment_cols = [col for col in df.columns if 'comment' in col.lower()]
    
    print(f"Found {len(comment_cols)} comment columns")
    
    # Process each case
    detailed_data = []
    integration_flows = defaultdict(lambda: defaultdict(list))
    integration_mappings = defaultdict(lambda: defaultdict(list))
    integration_errors = defaultdict(list)
    
    for idx, row in df.iterrows():
        case_key = row[key_col]
        case_type = row[case_type_col] if pd.notna(row[case_type_col]) else 'Unknown'
        integration = row[integration_col] if pd.notna(row[integration_col]) else 'N/A'
        summary = row[summary_col] if pd.notna(row[summary_col]) else ''
        description = row[description_col] if pd.notna(row[description_col]) else ''
        priority = row[priority_col] if pd.notna(row[priority_col]) else 'P3'
        status = row[status_col]
        resolution = row[resolution_col] if pd.notna(row[resolution_col]) else 'N/A'
        resolution_comments = row.get('Custom field (Resolution Comments)', '') if pd.notna(row.get('Custom field (Resolution Comments)', '')) else ''
        
        # Combine all text for analysis
        all_text = f"{summary}\n{description}"
        for comment_col in comment_cols:
            if pd.notna(row[comment_col]):
                all_text += f"\n{row[comment_col]}"
        
        # Extract detailed information
        flows = extract_flow_names(all_text)
        mappings = extract_field_mappings(all_text)
        errors = extract_error_messages(all_text)
        pre_prd = extract_pre_prd_references(all_text)
        record_types = extract_record_types(all_text)
        
        # Store data
        detailed_data.append({
            'Case Key': case_key,
            'Case Type': case_type,
            'Integration': integration,
            'Priority': priority,
            'Status': status,
            'Resolution': resolution,
            'Summary': summary[:200],
            'Resolution Comments': str(resolution_comments) if resolution_comments else '',
            'Flows Identified': ' | '.join(flows) if flows else 'Not specified',
            'Field/Mapping Issues': ' | '.join(mappings) if mappings else 'Not specified',
            'Error Messages': ' | '.join(errors[:3]) if errors else 'Not specified',
            'PRE/PRD References': ', '.join(pre_prd) if pre_prd else 'None',
            'Record Types': ', '.join(record_types) if record_types else 'Not specified',
            'Flow Count': len(flows),
            'Mapping Count': len(mappings),
            'Error Count': len(errors)
        })
        
        # Aggregate by integration
        if integration != 'N/A':
            for flow in flows:
                integration_flows[integration][flow].append(case_key)
            
            for mapping in mappings:
                integration_mappings[integration][mapping].append(case_key)
            
            for error in errors:
                integration_errors[integration].append({
                    'case': case_key,
                    'error': error,
                    'priority': priority
                })
    
    cases_df = pd.DataFrame(detailed_data)
    
    print(f"\n📊 Analysis Complete:")
    print(f"  Total flows identified: {cases_df['Flow Count'].sum()}")
    print(f"  Total mappings identified: {cases_df['Mapping Count'].sum()}")
    print(f"  Total errors extracted: {cases_df['Error Count'].sum()}")
    
    # ============================================================================
    # CREATE INTEGRATION-SPECIFIC SHEETS
    # ============================================================================
    
    # Enhanced Flow Analysis by Integration with more details
    flow_analysis = []
    for integration, flows in integration_flows.items():
        for flow, case_keys in flows.items():
            # Get details for these cases
            flow_cases = cases_df[cases_df['Case Key'].isin(case_keys)]
            
            # Determine flow direction and record type from flow name
            flow_lower = flow.lower()
            
            # Flow direction
            if 'to' in flow_lower:
                if 'netsuite' in flow_lower and flow_lower.index('netsuite') > flow_lower.index('to'):
                    direction = 'Import to NetSuite'
                elif 'netsuite' in flow_lower and flow_lower.index('netsuite') < flow_lower.index('to'):
                    direction = 'Export from NetSuite'
                else:
                    direction = 'Sync'
            elif 'import' in flow_lower:
                direction = 'Import'
            elif 'export' in flow_lower:
                direction = 'Export'
            elif 'sync' in flow_lower:
                direction = 'Sync'
            else:
                direction = 'Unspecified'
            
            # Record type
            record_type = 'N/A'
            if 'sales order' in flow_lower or 'order' in flow_lower:
                record_type = 'Sales Order'
            elif 'cash sale' in flow_lower:
                record_type = 'Cash Sale'
            elif 'fulfillment' in flow_lower or 'item fulfillment' in flow_lower:
                record_type = 'Item Fulfillment'
            elif 'refund' in flow_lower or 'credit memo' in flow_lower:
                record_type = 'Refund/Credit'
            elif 'settlement' in flow_lower:
                record_type = 'Settlement'
            elif 'shipment' in flow_lower:
                record_type = 'Shipment'
            elif 'customer' in flow_lower:
                record_type = 'Customer'
            elif 'product' in flow_lower or 'item' in flow_lower or 'inventory' in flow_lower:
                record_type = 'Product/Item'
            elif 'payment' in flow_lower:
                record_type = 'Payment'
            elif 'invoice' in flow_lower:
                record_type = 'Invoice'
            
            # Priority breakdown
            priority_counts = flow_cases['Priority'].value_counts()
            p1_count = priority_counts.get('P1', 0)
            p2_count = priority_counts.get('P2', 0)
            
            # Status breakdown
            status_counts = flow_cases['Status'].str.lower().apply(lambda x: 'Open' if x not in ['closed', 'resolved'] else 'Closed').value_counts()
            open_count = status_counts.get('Open', 0)
            closed_count = status_counts.get('Closed', 0)
            
            # Common errors for this flow
            all_errors = []
            for err in flow_cases['Error Messages'].dropna():
                if str(err) != 'N/A' and str(err) != 'Not specified':
                    all_errors.append(str(err)[:80])
            common_error = all_errors[0] if all_errors else 'N/A'
            
            # PRE/PRD references
            pre_refs = []
            for ref in flow_cases['PRE/PRD References'].dropna():
                if str(ref) != 'N/A' and str(ref) != 'Not specified':
                    pre_refs.extend(str(ref).split(', '))
            unique_pres = list(set(pre_refs))[:3]
            pre_summary = ', '.join(unique_pres) if unique_pres else 'N/A'
            
            flow_analysis.append({
                'Integration': integration,
                'Flow Name': flow,
                'Direction': direction,
                'Record Type': record_type,
                'Issue Count': len(case_keys),
                'Open': open_count,
                'Closed': closed_count,
                'P1': p1_count,
                'P2': p2_count,
                'Affected Cases': ', '.join(case_keys[:10]),
                'Common Error': common_error,
                'PRE/PRD Refs': pre_summary,
                'Priority': 'Critical' if p1_count > 0 else 'High' if len(case_keys) > 5 else 'Medium'
            })
    
    flows_df = pd.DataFrame(flow_analysis).sort_values(['Integration', 'Issue Count'], ascending=[True, False])
    
    # Mapping Analysis by Integration
    mapping_analysis = []
    for integration, mappings in integration_mappings.items():
        for mapping, cases in mappings.items():
            mapping_analysis.append({
                'Integration': integration,
                'Field/Mapping': mapping,
                'Issue Count': len(cases),
                'Affected Cases': ', '.join(cases[:10]),
                'Sample Case': cases[0] if cases else '',
                'Priority': 'High' if len(cases) > 2 else 'Medium'
            })
    
    mappings_df = pd.DataFrame(mapping_analysis).sort_values(['Integration', 'Issue Count'], ascending=[True, False])
    
    # Error Analysis by Integration
    error_analysis = []
    for integration, errors in integration_errors.items():
        error_groups = defaultdict(list)
        for error_info in errors:
            error_groups[error_info['error']].append(error_info['case'])
        
        for error, cases in error_groups.items():
            error_analysis.append({
                'Integration': integration,
                'Error Message': error[:200],
                'Occurrence Count': len(cases),
                'Affected Cases': ', '.join(cases[:10]),
                'Sample Case': cases[0] if cases else ''
            })
    
    errors_df = pd.DataFrame(error_analysis).sort_values(['Integration', 'Occurrence Count'], ascending=[True, False])
    
    # Helper function to check if flow/mapping is meaningful
    def is_meaningful_data(text):
        """Check if this is meaningful data, not metadata."""
        if not text or text == 'N/A' or pd.isna(text):
            return False
        text_lower = str(text).lower()
        
        # Filter out metadata
        exclude_patterns = ['accountid', 'yes/no', 'link to video', 'file name', 
                          'slack', 'h2.io', 'mailto:', '@celigo.com']
        if any(pattern in text_lower for pattern in exclude_patterns):
            return False
        
        # Must contain flow keywords
        flow_keywords = ['order', 'sales', 'import', 'export', 'sync', 'fulfillment',
                        'settlement', 'shipment', 'refund', 'payment', 'invoice',
                        'customer', 'product', 'item', 'inventory', 'cash sale', 'field',
                        'credit memo', 'journal', 'mapping', 'record', 'data']
        return any(keyword in text_lower for keyword in flow_keywords)
    
    # Helper function to normalize flow names for merging
    def normalize_flow_name(flow_name):
        """Normalize flow names to merge similar variations."""
        if not flow_name or pd.isna(flow_name):
            return ''
        
        flow = str(flow_name).lower().strip()
        
        # Remove common prefixes/suffixes
        flow = flow.replace('the ', '').replace('a ', '').replace('an ', '')
        flow = flow.replace(' flow', '').replace(' sync', '').replace(' import', '').replace(' export', '')
        
        # Normalize common variations
        replacements = {
            'salesorder': 'sales order',
            'purchaseorder': 'purchase order',
            'itemfulfillment': 'item fulfillment',
            'cashsale': 'cash sale',
            'creditmemo': 'credit memo',
            'customerdeposit': 'customer deposit',
            'customerpayment': 'customer payment',
            'journalentry': 'journal entry',
            'to netsuite': '→ netsuite',
            'from netsuite': 'netsuite →',
            'to ns': '→ netsuite',
            'from ns': 'netsuite →',
            'sf to': 'salesforce →',
            'to sf': '→ salesforce',
            'from sf': 'salesforce →',
            'shopify to': 'shopify →',
            'to shopify': '→ shopify',
            'amazon to': 'amazon →',
            'to amazon': '→ amazon',
            'bigcommerce to': 'bigcommerce →',
            'to bigcommerce': '→ bigcommerce'
        }
        
        for old, new in replacements.items():
            flow = flow.replace(old, new)
        
        # Remove extra whitespace
        flow = ' '.join(flow.split())
        
        return flow
    
    # Function to merge similar flows
    def merge_similar_flows(flows_df):
        """Merge flows with similar normalized names."""
        if flows_df.empty:
            return flows_df
        
        # Create normalized names for grouping
        flows_df['Normalized'] = flows_df['Flow Name'].apply(normalize_flow_name)
        
        # Group by Integration and Normalized name
        merged_flows = []
        grouped = flows_df.groupby(['Integration', 'Normalized'])
        
        for (integration, norm_name), group in grouped:
            if not norm_name:  # Skip empty normalized names
                continue
                
            # Use the most common original flow name as the representative
            original_names = group['Flow Name'].value_counts()
            representative_name = original_names.index[0]
            
            # Merge the data
            all_cases = []
            for cases_str in group['Affected Cases']:
                if pd.notna(cases_str):
                    all_cases.extend(cases_str.split(', '))
            unique_cases = list(set(all_cases))[:10]
            
            # Aggregate PRE/PRD refs
            all_refs = []
            for refs in group['PRE/PRD Refs']:
                if pd.notna(refs) and refs != 'N/A':
                    all_refs.extend(refs.split(', '))
            unique_refs = list(set(all_refs))[:3]
            
            # Get most common error
            errors = group['Common Error'].dropna()
            common_error = errors.mode()[0] if len(errors) > 0 and not errors.mode().empty else group['Common Error'].iloc[0]
            
            merged_flows.append({
                'Integration': integration,
                'Flow Name': representative_name,
                'Direction': group['Direction'].iloc[0],
                'Record Type': group['Record Type'].iloc[0],
                'Issue Count': group['Issue Count'].sum(),
                'Open': group['Open'].sum(),
                'Closed': group['Closed'].sum(),
                'P1': group['P1'].sum(),
                'P2': group['P2'].sum(),
                'Affected Cases': ', '.join(unique_cases),
                'Common Error': common_error,
                'PRE/PRD Refs': ', '.join(unique_refs) if unique_refs else 'N/A',
                'Priority': 'Critical' if group['P1'].sum() > 0 else 'High' if group['Issue Count'].sum() > 5 else 'Medium',
                'Merged Count': len(group)  # Track how many variations were merged
            })
        
        result_df = pd.DataFrame(merged_flows).sort_values(['Integration', 'Issue Count'], ascending=[True, False])
        return result_df
    
    # Integration Summary with improved metrics
    integration_summary = []
    for integration in cases_df[cases_df['Integration'] != 'N/A']['Integration'].unique():
        int_cases = cases_df[cases_df['Integration'] == integration]
        int_flows = integration_flows.get(integration, {})
        int_mappings = integration_mappings.get(integration, {})
        int_errors = integration_errors.get(integration, [])
        
        total_cases = len(int_cases)
        closed_cases = len(int_cases[int_cases['Status'].str.lower().isin(['closed', 'resolved'])])
        open_cases = total_cases - closed_cases
        p1_cases = len(int_cases[int_cases['Priority'] == 'P1'])
        p1_open = len(int_cases[(int_cases['Priority'] == 'P1') & (~int_cases['Status'].str.lower().isin(['closed', 'resolved']))])
        
        # Get meaningful top flow
        meaningful_flows = {f: cases for f, cases in int_flows.items() if is_meaningful_data(f)}
        top_flow = max(meaningful_flows.items(), key=lambda x: len(x[1]))[0] if meaningful_flows else 'N/A'
        top_flow_count = len(meaningful_flows[top_flow]) if top_flow != 'N/A' else 0
        
        # Get meaningful top error
        unique_errors = list(set([e['error'] for e in int_errors]))
        meaningful_errors = [err for err in unique_errors if is_meaningful_data(err)]
        top_error = meaningful_errors[0][:80] if meaningful_errors else 'N/A'
        
        # Count frequent flows (2+ occurrences) for more actionable metric
        frequent_flows = {f: cases for f, cases in meaningful_flows.items() if len(cases) >= 2}
        
        integration_summary.append({
            'Integration': integration,
            'Total Cases': total_cases,
            'Open': open_cases,
            'Closed': closed_cases,
            'Resolution Rate': f"{(closed_cases/total_cases*100):.0f}%" if total_cases > 0 else '0%',
            'P1 Total': p1_cases,
            'P1 Open': p1_open,
            'Frequent Flows (2+)': len(frequent_flows),
            'Unique Errors': len(meaningful_errors),
            'Top Flow Issue': f"{top_flow} ({top_flow_count} cases)" if top_flow != 'N/A' else 'N/A',
            'Most Common Error': top_error
        })
    
    # Sort by impact score (total cases + P1 weight)
    summary_df = pd.DataFrame(integration_summary)
    summary_df['_impact_score'] = summary_df['Total Cases'] + (summary_df['P1 Total'] * 10)
    summary_df = summary_df.sort_values('_impact_score', ascending=False).drop('_impact_score', axis=1)
    
    # ============================================================================
    # CREATE CURATED ACTIONABLE REPORT (ONLY ONE FILE NEEDED)
    # ============================================================================
    
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'2025_Actionable_Deep_Dive.xlsx'
    
    print(f"\n📝 Creating actionable deep dive report: {output_file}")
    
    # Filter for actionable insights (2+ occurrences)
    clean_flows = flows_df[flows_df['Flow Name'].apply(is_meaningful_data)]
    frequent_flows_raw = clean_flows[clean_flows['Issue Count'] >= 2]
    
    # Merge similar/same flow names
    frequent_flows = merge_similar_flows(frequent_flows_raw)
    
    # Remove the 'Normalized' column if it exists (used internally for merging)
    if 'Normalized' in frequent_flows.columns:
        frequent_flows = frequent_flows.drop(columns=['Normalized'])
    
    frequent_errors = errors_df[errors_df['Occurrence Count'] >= 2].sort_values(['Integration', 'Occurrence Count'], ascending=[True, False])
    
    # ============================================================================
    # CREATE ADDITIONAL SUMMARY SHEETS
    # ============================================================================
    
    # Count by Integration App (all case types)
    integration_counts = []
    for integration in cases_df[cases_df['Integration'] != 'N/A']['Integration'].unique():
        int_cases = cases_df[cases_df['Integration'] == integration]
        
        # Case type breakdown
        case_type_counts = int_cases['Case Type'].value_counts()
        
        # Priority breakdown
        priority_counts = int_cases['Priority'].value_counts()
        
        # Status breakdown
        status_counts = int_cases['Status'].str.lower().apply(lambda x: 'Open' if x not in ['closed', 'resolved'] else 'Closed').value_counts()
        
        integration_counts.append({
            'Integration': integration,
            'Total Cases': len(int_cases),
            'Bug': case_type_counts.get('Bug', 0),
            'Query': case_type_counts.get('Query', 0),
            'Documentation': case_type_counts.get('Documentation', 0),
            'Product Enhancement': case_type_counts.get('Product Enhancement', 0),
            'Open': status_counts.get('Open', 0),
            'Closed': status_counts.get('Closed', 0),
            'P1': priority_counts.get('P1', 0),
            'P2': priority_counts.get('P2', 0),
            'P3': priority_counts.get('P3', 0),
            'P4': priority_counts.get('P4', 0)
        })
    
    integration_count_df = pd.DataFrame(integration_counts).sort_values('Total Cases', ascending=False)
    
    # ============================================================================
    # RESOLUTION BREAKDOWN BY INTEGRATION APP
    # ============================================================================
    
    # Create detailed resolution breakdown for each integration
    resolution_breakdown = []
    for integration in cases_df[cases_df['Integration'] != 'N/A']['Integration'].unique():
        int_cases = cases_df[cases_df['Integration'] == integration]
        
        # Get all resolution counts
        resolution_counts = int_cases['Resolution'].value_counts()
        
        # Create row with integration name and all resolutions
        row_data = {'Integration': integration, 'Total Cases': len(int_cases)}
        
        # Add each resolution as a separate column
        for resolution, count in resolution_counts.items():
            if pd.notna(resolution):
                row_data[resolution] = count
        
        resolution_breakdown.append(row_data)
    
    # Create DataFrame and fill NaN with 0
    resolution_breakdown_df = pd.DataFrame(resolution_breakdown).fillna(0)
    
    # Sort by Total Cases descending
    resolution_breakdown_df = resolution_breakdown_df.sort_values('Total Cases', ascending=False)
    
    # Convert resolution columns to int
    for col in resolution_breakdown_df.columns:
        if col not in ['Integration', 'Total Cases']:
            resolution_breakdown_df[col] = resolution_breakdown_df[col].astype(int)
    
    # ============================================================================
    # CUSTOMER-SPECIFIC ANALYSIS
    # ============================================================================
    
    # Find customer and tier columns
    customer_col = None
    tier_col = None
    severity_col = None
    
    for col in df.columns:
        if 'customer' in col.lower() and 'old' in col.lower():
            customer_col = col
        elif 'customer tier' in col.lower():
            tier_col = col
        elif col.lower() == 'severity':
            severity_col = col
        elif 'severity' in col.lower() and 'custom field' in col.lower():
            severity_col = col
    
    customer_analysis = []
    if customer_col and pd.notna(df[customer_col]).sum() > 0:
        # Get resolution column
        resolution_col = None
        for col in df.columns:
            if col.lower() == 'resolution':
                resolution_col = col
                break
        
        # Get unique customers (including '- None -' but we'll mark it)
        for customer in df[customer_col].dropna().unique():
            if customer and str(customer).strip() and str(customer) != 'N/A':
                cust_cases = cases_df[cases_df['Case Key'].isin(
                    df[df[customer_col] == customer]['Issue key'].values
                )]
                
                if len(cust_cases) == 0:
                    continue
                
                # Mark if internal
                is_internal = '✓' if str(customer) == '- None -' else ''
                
                # Get tier for this customer
                cust_tier = df[df[customer_col] == customer][tier_col].dropna().mode()
                tier = cust_tier.values[0] if len(cust_tier) > 0 else 'N/A'
                
                # Severity breakdown
                severity_counts = df[df[customer_col] == customer][severity_col].value_counts()
                
                # Case type breakdown
                case_type_counts = cust_cases['Case Type'].value_counts()
                
                # Priority breakdown
                priority_counts = cust_cases['Priority'].value_counts()
                
                # Status
                status_counts = cust_cases['Status'].str.lower().apply(
                    lambda x: 'Open' if x not in ['closed', 'resolved'] else 'Closed'
                ).value_counts()
                
                # Top integration for this customer
                top_int = cust_cases['Integration'].value_counts()
                top_integration = top_int.index[0] if len(top_int) > 0 else 'N/A'
                
                # Resolution breakdown (if available)
                resolution_types = []
                if resolution_col:
                    for case_key in df[df[customer_col] == customer]['Issue key'].values:
                        res = df[df['Issue key'] == case_key][resolution_col].values
                        if len(res) > 0 and pd.notna(res[0]):
                            resolution_types.append(str(res[0]))
                
                top_resolution = pd.Series(resolution_types).mode()
                resolution = top_resolution.values[0] if len(top_resolution) > 0 else 'N/A'
                
                # Customer health score (weighted: P1*10 + P2*5 + S1*10 + Open*3)
                health_score = (priority_counts.get('P1', 0) * 10 + 
                              priority_counts.get('P2', 0) * 5 + 
                              severity_counts.get('S1', 0) * 10 + 
                              status_counts.get('Open', 0) * 3)
                
                customer_analysis.append({
                    'Customer': str(customer)[:50],
                    'Internal': is_internal,
                    'Tier': tier,
                    'Total Cases': len(cust_cases),
                    'Open': status_counts.get('Open', 0),
                    'Closed': status_counts.get('Closed', 0),
                    'Bug': case_type_counts.get('Bug', 0),
                    'Query': case_type_counts.get('Query', 0),
                    'Top Resolution': resolution[:40] if resolution != 'N/A' else 'N/A',
                    'Top Integration': top_integration
                })
        
        customer_analysis_df = pd.DataFrame(customer_analysis).sort_values('Total Cases', ascending=False)
    else:
        customer_analysis_df = pd.DataFrame([{'Customer': 'N/A', 'Note': 'No customer data available'}])
    
    # ============================================================================
    # ERROR CATEGORY ANALYSIS
    # ============================================================================
    
    def categorize_error(error_text):
        """Categorize error into specific categories - IMPROVED with more granular categories."""
        if pd.isna(error_text) or str(error_text) in ['N/A', 'Not specified']:
            return 'Unspecified'
        
        error_lower = str(error_text).lower()
        
        # Hook/Script errors (check first - very specific)
        if any(x in error_lower for x in ['hook error', 'hook function', 'script error', 'nlobjsearch', 'customscript', 'scriptid']):
            return 'Hook/Script Error'
        
        # Kit/BOM errors
        elif any(x in error_lower for x in ['kit definition', 'bom', 'kit component', 'member item']):
            return 'Kit/BOM Issue'
        
        # Storemap errors
        elif any(x in error_lower for x in ['storemap', 'store map', 'missing storemap']):
            return 'Storemap Issue'
        
        # Integration App errors
        elif any(x in error_lower for x in ['integration app', 'cannot delete a resource that belongs to', 'ia deleted', 'ia error']):
            return 'Integration App Error'
        
        # Sublist operation errors (NetSuite specific)
        elif any(x in error_lower for x in ['sublist', 'invalid sublist', 'sublist operation', 'line item']):
            return 'Sublist Operation'
        
        # Search/Query errors (NetSuite)
        elif any(x in error_lower for x in ['search', 'searchid', 'unable to get export searchid', 'invalid search']):
            return 'Search/Query Error'
        
        # Webhook errors
        elif any(x in error_lower for x in ['webhook', 'web hook']):
            return 'Webhook Error'
        
        # Authentication errors
        elif any(x in error_lower for x in ['401', '403', 'unauthorized', 'authentication', 'token', 'auth', 'jwt', 'credential', 'reauthenticate']):
            return 'Authentication'
        
        # Mapping/Field errors
        elif any(x in error_lower for x in ['mapping error', 'field error', 'missing field', 'invalid field', 'invalid column']):
            return 'Mapping/Field'
        
        # Record errors
        elif any(x in error_lower for x in ['failed to create', 'failed to update', 'failed to save', 'failed to add', 'cannot create', 'unable to create']):
            return 'Record Creation/Update'
        
        # Rate limit/Performance
        elif any(x in error_lower for x in ['rate limit', 'too many requests', '429', 'performance', 'slow']):
            return 'Rate Limit/Performance'
        
        # Data validation
        elif any(x in error_lower for x in ['validation', 'invalid value', 'invalid format', 'required', 'must be', 'must enter']):
            return 'Data Validation'
        
        # Network/Connection
        elif any(x in error_lower for x in ['connection', 'network', '502', '503', '504', 'unreachable', 'timeout', 'timed out']):
            return 'Network/Connection'
        
        # API errors
        elif any(x in error_lower for x in ['api error', '400', '404', '500', 'bad request', 'status code']):
            return 'API Error'
        
        # Configuration/Setup
        elif any(x in error_lower for x in ['config', 'setup', 'install', 'uninstall', 'not configured', 'missing connector']):
            return 'Configuration/Setup'
        
        # File/Bundle errors
        elif any(x in error_lower for x in ['failed to load file', 'bundle', 'file size', 'suitebundles']):
            return 'File/Bundle Error'
        
        else:
            return 'Other'
    
    error_category_analysis = []
    for idx, row in cases_df.iterrows():
        errors = row['Error Messages']
        if pd.notna(errors) and str(errors) not in ['N/A', 'Not specified']:
            error_list = str(errors).split('|')
            for error in error_list[:3]:  # Top 3 errors per case
                category = categorize_error(error)
                error_category_analysis.append({
                    'Category': category,
                    'Error': error.strip()[:100],
                    'Integration': row['Integration'],
                    'Case Key': row['Case Key'],
                    'Priority': row['Priority'],
                    'Status': row['Status']
                })
    
    error_cat_df = pd.DataFrame(error_category_analysis)
    
    # Summarize by category
    if len(error_cat_df) > 0:
        error_summary = []
        for category in error_cat_df['Category'].unique():
            cat_errors = error_cat_df[error_cat_df['Category'] == category]
            
            # Status breakdown
            status_counts = cat_errors['Status'].str.lower().apply(
                lambda x: 'Open' if x not in ['closed', 'resolved'] else 'Closed'
            ).value_counts()
            
            # Priority breakdown
            priority_counts = cat_errors['Priority'].value_counts()
            
            # Top integration
            top_int = cat_errors['Integration'].value_counts()
            top_integration = f"{top_int.index[0]}({top_int.values[0]})" if len(top_int) > 0 else 'N/A'
            
            # Most common error in this category
            error_counts = cat_errors['Error'].value_counts()
            common_error = error_counts.index[0] if len(error_counts) > 0 else 'N/A'
            
            error_summary.append({
                'Error Category': category,
                'Total Occurrences': len(cat_errors),
                'Unique Errors': cat_errors['Error'].nunique(),
                'Open': status_counts.get('Open', 0),
                'Closed': status_counts.get('Closed', 0),
                'P1': priority_counts.get('P1', 0),
                'P2': priority_counts.get('P2', 0),
                'Top Integration': top_integration,
                'Most Common Error': common_error[:80]
            })
        
        error_category_summary_df = pd.DataFrame(error_summary).sort_values('Total Occurrences', ascending=False)
    else:
        error_category_summary_df = pd.DataFrame([{'Error Category': 'N/A', 'Note': 'No errors extracted'}])
    
    # ============================================================================
    # BUG QUALITY ANALYSIS - Resolution Types
    # ============================================================================
    
    # Find Bug Resolution column
    bug_resolution_col = None
    for col in df.columns:
        if 'bug resolution' in col.lower():
            bug_resolution_col = col
            break
    
    # Find Skip QA column
    skip_qa_col = None
    for col in df.columns:
        if 'skip qa' in col.lower():
            skip_qa_col = col
            break
    
    # Find Assignee column
    assignee_col = None
    for col in df.columns:
        if col == 'Assignee':
            assignee_col = col
            break
    
    bug_quality_analysis = []
    if bug_resolution_col and pd.notna(df[bug_resolution_col]).sum() > 0:
        # Overall resolution breakdown
        resolution_counts = df[bug_resolution_col].value_counts()
        total_resolutions = resolution_counts.sum()
        
        for resolution, count in resolution_counts.items():
            if resolution and str(resolution).lower() not in ['nan', 'n/a', '']:
                pct = (count / total_resolutions) * 100
                
                # Get affected cases
                resolution_cases = df[df[bug_resolution_col] == resolution]
                
                # Case type breakdown
                case_type_counts = resolution_cases['Case Type[Dropdown]'].value_counts() if 'Case Type[Dropdown]' in resolution_cases.columns else {}
                bug_count = case_type_counts.get('Bug', 0)
                
                # Priority breakdown
                priority_counts = resolution_cases['Priority'].value_counts()
                
                # Status breakdown
                status_counts = resolution_cases['Status'].str.lower().apply(
                    lambda x: 'Open' if x not in ['closed', 'resolved'] else 'Closed'
                ).value_counts()
                
                # Top integration for this resolution type
                int_col = None
                for col in resolution_cases.columns:
                    if 'integration' in col.lower() and 'dropdown' in col.lower():
                        int_col = col
                        break
                
                top_integration = 'N/A'
                if int_col:
                    int_counts = resolution_cases[int_col].value_counts()
                    if len(int_counts) > 0:
                        top_integration = f"{int_counts.index[0]} ({int_counts.values[0]})"
                
                bug_quality_analysis.append({
                    'Resolution Type': str(resolution),
                    'Total Cases': count,
                    'Percentage': f"{pct:.1f}%",
                    'Bug Cases': bug_count,
                    'Open': status_counts.get('Open', 0),
                    'Closed': status_counts.get('Closed', 0),
                    'P1': priority_counts.get('P1', 0),
                    'P2': priority_counts.get('P2', 0),
                    'Top Integration': top_integration
                })
        
        bug_quality_df = pd.DataFrame(bug_quality_analysis).sort_values('Total Cases', ascending=False)
        
        # Add QA Coverage Analysis
        qa_coverage_analysis = []
        if skip_qa_col:
            # Overall QA coverage
            skip_qa_counts = df[skip_qa_col].value_counts()
            total_qa = skip_qa_counts.sum()
            
            for skip_value, count in skip_qa_counts.items():
                if skip_value and str(skip_value).lower() not in ['nan', 'n/a', '']:
                    pct = (count / total_qa) * 100
                    
                    qa_cases = df[df[skip_qa_col] == skip_value]
                    
                    # Bug count for this QA status
                    case_type_counts = qa_cases['Case Type[Dropdown]'].value_counts() if 'Case Type[Dropdown]' in qa_cases.columns else {}
                    bug_count = case_type_counts.get('Bug', 0)
                    
                    # Bug resolution breakdown
                    bug_resolutions = []
                    if bug_resolution_col:
                        bug_cases = qa_cases[qa_cases['Case Type[Dropdown]'] == 'Bug'] if 'Case Type[Dropdown]' in qa_cases.columns else pd.DataFrame()
                        if len(bug_cases) > 0:
                            res_counts = bug_cases[bug_resolution_col].value_counts()
                            bug_resolutions = [f"{k}({v})" for k, v in res_counts.head(3).items()]
                    
                    qa_coverage_analysis.append({
                        'QA Status': f"Skip QA: {skip_value}",
                        'Total Cases': count,
                        'Percentage': f"{pct:.1f}%",
                        'Bugs Found': bug_count,
                        'Bug Rate': f"{(bug_count/count*100):.1f}%" if count > 0 else '0%',
                        'Top Bug Resolutions': ', '.join(bug_resolutions) if bug_resolutions else 'N/A'
                    })
            
            qa_coverage_df = pd.DataFrame(qa_coverage_analysis).sort_values('Total Cases', ascending=False)
        else:
            qa_coverage_df = pd.DataFrame([{'QA Status': 'N/A', 'Note': 'No QA data available'}])
        
        # Resolution by Integration
        resolution_by_ia = []
        int_col = None
        for col in df.columns:
            if 'integration' in col.lower() and 'dropdown' in col.lower():
                int_col = col
                break
        
        if int_col and bug_resolution_col:
            for integration in df[df[int_col].notna()][int_col].unique():
                if str(integration) not in ['N/A', 'nan', '']:
                    int_cases = df[df[int_col] == integration]
                    
                    # Resolution breakdown
                    res_counts = int_cases[bug_resolution_col].value_counts()
                    total_res = res_counts.sum()
                    
                    if total_res > 0:
                        code_fix = res_counts.get('Code fix', 0)
                        config = res_counts.get('Configuration', 0)
                        other = res_counts.get('Other', 0)
                        
                        resolution_by_ia.append({
                            'Integration': str(integration),
                            'Total with Resolution': int(total_res),
                            'Code Fix': int(code_fix),
                            'Code Fix %': f"{(code_fix/total_res*100):.1f}%",
                            'Configuration': int(config),
                            'Config %': f"{(config/total_res*100):.1f}%",
                            'Other': int(other),
                            'Other %': f"{(other/total_res*100):.1f}%",
                            'Quality Score': f"{((code_fix/total_res)*100):.0f}"  # Higher = more code issues
                        })
            
            resolution_by_ia_df = pd.DataFrame(resolution_by_ia).sort_values('Code Fix', ascending=False)
        else:
            resolution_by_ia_df = pd.DataFrame([{'Integration': 'N/A', 'Note': 'No resolution data by IA'}])
        
    else:
        bug_quality_df = pd.DataFrame([{'Resolution Type': 'N/A', 'Note': 'No bug resolution data available'}])
        qa_coverage_df = pd.DataFrame([{'QA Status': 'N/A', 'Note': 'No QA data available'}])
        resolution_by_ia_df = pd.DataFrame([{'Integration': 'N/A', 'Note': 'No resolution data by IA'}])
    
    # ============================================================================
    # DETAILED ERROR DISTRIBUTION BY INTEGRATION (NEW)
    # ============================================================================
    
    # Create detailed breakdown: Error Category x Integration
    error_distribution = []
    if len(error_cat_df) > 0:
        for category in sorted(error_cat_df['Category'].unique()):
            cat_errors = error_cat_df[error_cat_df['Category'] == category]
            
            # Count by integration
            integration_counts = cat_errors.groupby('Integration').size().sort_values(ascending=False)
            
            # Get status and priority breakdown per integration
            for integration, count in integration_counts.items():
                int_cat_errors = cat_errors[cat_errors['Integration'] == integration]
                
                # Status breakdown
                status_counts = int_cat_errors['Status'].str.lower().apply(
                    lambda x: 'Open' if x not in ['closed', 'resolved'] else 'Closed'
                ).value_counts()
                
                # Priority breakdown
                priority_counts = int_cat_errors['Priority'].value_counts()
                
                # Get sample errors
                error_samples = int_cat_errors['Error'].unique()[:2]
                sample_error = error_samples[0] if len(error_samples) > 0 else 'N/A'
                
                # Get affected cases
                affected_cases = ', '.join(int_cat_errors['Case Key'].unique()[:5])
                
                error_distribution.append({
                    'Error Category': category,
                    'Integration': integration,
                    'Error Count': count,
                    'Open': status_counts.get('Open', 0),
                    'Closed': status_counts.get('Closed', 0),
                    'P1': priority_counts.get('P1', 0),
                    'P2': priority_counts.get('P2', 0),
                    'Sample Error': sample_error[:80],
                    'Affected Cases': affected_cases
                })
        
        error_distribution_df = pd.DataFrame(error_distribution).sort_values(
            ['Error Category', 'Error Count'], ascending=[True, False]
        )
    else:
        error_distribution_df = pd.DataFrame([{'Error Category': 'N/A', 'Note': 'No errors extracted'}])
    
    # Count by Month
    # Extract month from Created date (assuming there's a Created or timestamp column)
    created_col = None
    for col in ['Created', 'Created Date', 'created', 'Timestamp']:
        if col in df.columns:
            created_col = col
            break
    
    if created_col:
        # Parse dates and extract month
        cases_df['Month'] = pd.to_datetime(df[created_col], errors='coerce').dt.to_period('M').astype(str)
        
        monthly_counts = []
        for month in sorted(cases_df['Month'].dropna().unique()):
            month_cases = cases_df[cases_df['Month'] == month]
            
            # Case type breakdown
            case_type_counts = month_cases['Case Type'].value_counts()
            
            # Priority breakdown
            priority_counts = month_cases['Priority'].value_counts()
            
            # Status breakdown
            status_counts = month_cases['Status'].str.lower().apply(lambda x: 'Open' if x not in ['closed', 'resolved'] else 'Closed').value_counts()
            
            # Top integrations for this month
            top_integrations = month_cases['Integration'].value_counts().head(3)
            top_int_str = ', '.join([f"{k}({v})" for k, v in top_integrations.items()])
            
            monthly_counts.append({
                'Month': month,
                'Total Cases': len(month_cases),
                'Bug': case_type_counts.get('Bug', 0),
                'Query': case_type_counts.get('Query', 0),
                'Documentation': case_type_counts.get('Documentation', 0),
                'Product Enhancement': case_type_counts.get('Product Enhancement', 0),
                'Open': status_counts.get('Open', 0),
                'Closed': status_counts.get('Closed', 0),
                'P1': priority_counts.get('P1', 0),
                'P2': priority_counts.get('P2', 0),
                'P3': priority_counts.get('P3', 0),
                'P4': priority_counts.get('P4', 0),
                'Top Integrations': top_int_str
            })
        
        monthly_count_df = pd.DataFrame(monthly_counts)
    else:
        # If no date column, create empty dataframe
        monthly_count_df = pd.DataFrame([{'Month': 'N/A', 'Total Cases': len(cases_df), 'Note': 'No date column found'}])
    
    # Write to Excel with formatting
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Integration Overview', index=False)
        integration_count_df.to_excel(writer, sheet_name='Count by Integration', index=False)
        resolution_breakdown_df.to_excel(writer, sheet_name='Resolution Breakdown by IA', index=False)
        monthly_count_df.to_excel(writer, sheet_name='Count by Month', index=False)
        customer_analysis_df.to_excel(writer, sheet_name='Customer Analysis', index=False)
        error_category_summary_df.to_excel(writer, sheet_name='Error Categories', index=False)
        error_distribution_df.to_excel(writer, sheet_name='Error Distribution by IA', index=False)
        frequent_flows.to_excel(writer, sheet_name='Frequent Flow Issues', index=False)
        frequent_errors.to_excel(writer, sheet_name='Recurring Errors', index=False)
        cases_df.to_excel(writer, sheet_name='Case Details', index=False)
        
        # Apply visual enhancements to all sheets
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Color scheme
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')  # Dark blue
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        # Priority/Status colors
        p1_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Light red
        p2_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')  # Light orange
        open_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Light yellow
        closed_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
        
        # Alternating row colors
        alt_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Light gray
        
        # Border style
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            # Auto-adjust column widths
            for idx, col in enumerate(worksheet.columns, 1):
                max_length = 0
                column = get_column_letter(idx)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 80)
                worksheet.column_dimensions[column].width = adjusted_width
            
            # Style header row
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Set row height for header
            worksheet.row_dimensions[1].height = 30
            
            # Apply alternating row colors and borders
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                # Alternating row background
                if row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = alt_row_fill
                
                # Apply borders to all cells
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Sheet-specific conditional formatting
            if sheet_name == 'Integration Overview':
                # Highlight P1 Open column
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    p1_open_col = None
                    for idx, cell in enumerate(worksheet[1], 1):
                        if cell.value == 'P1 Open':
                            p1_open_col = idx
                            break
                    if p1_open_col and row[p1_open_col-1].value and int(row[p1_open_col-1].value) > 0:
                        row[p1_open_col-1].fill = p1_fill
                        row[p1_open_col-1].font = Font(bold=True)
            
            elif sheet_name == 'Customer Analysis':
                # Highlight internal customers
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    internal_col = None
                    for idx, cell in enumerate(worksheet[1], 1):
                        if cell.value == 'Internal':
                            internal_col = idx
                            break
                    if internal_col and row[internal_col-1].value == '✓':
                        for cell in row:
                            if cell.fill.start_color.rgb in ['F2F2F2', '00000000']:
                                cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            
            elif sheet_name == 'Error Categories':
                # Color code by severity
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    p1_col = None
                    for idx, cell in enumerate(worksheet[1], 1):
                        if cell.value == 'P1':
                            p1_col = idx
                            break
                    if p1_col and row[p1_col-1].value and int(row[p1_col-1].value) > 0:
                        row[0].font = Font(bold=True, color='C00000')  # Red for categories with P1
            
            elif sheet_name == 'Error Distribution by IA':
                # Highlight high error counts
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    error_count_col = None
                    for idx, cell in enumerate(worksheet[1], 1):
                        if cell.value == 'Error Count':
                            error_count_col = idx
                            break
                    if error_count_col and row[error_count_col-1].value:
                        count = int(row[error_count_col-1].value)
                        if count >= 5:
                            row[error_count_col-1].fill = p1_fill
                            row[error_count_col-1].font = Font(bold=True)
                        elif count >= 3:
                            row[error_count_col-1].fill = p2_fill
            
            elif sheet_name == 'Frequent Flow Issues':
                # Highlight by priority
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    priority_col = None
                    for idx, cell in enumerate(worksheet[1], 1):
                        if cell.value == 'Priority':
                            priority_col = idx
                            break
                    if priority_col:
                        if row[priority_col-1].value == 'Critical':
                            row[priority_col-1].fill = p1_fill
                            row[priority_col-1].font = Font(bold=True)
                        elif row[priority_col-1].value == 'High':
                            row[priority_col-1].fill = p2_fill
            
            elif sheet_name == 'Case Details':
                # Highlight open P1 cases
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    priority_col = status_col = None
                    for idx, cell in enumerate(worksheet[1], 1):
                        if cell.value == 'Priority':
                            priority_col = idx
                        elif cell.value == 'Status':
                            status_col = idx
                    
                    if priority_col and status_col:
                        is_p1 = row[priority_col-1].value == 'P1'
                        is_open = str(row[status_col-1].value).lower() not in ['closed', 'resolved']
                        
                        if is_p1 and is_open:
                            row[0].fill = p1_fill  # Highlight case key
                            row[0].font = Font(bold=True)
            
            # Freeze panes (first row and first column)
            worksheet.freeze_panes = 'B2'
    
    print(f"\n✅ Actionable deep dive report complete!")
    print(f"\n📋 File contains 10 comprehensive sheets:")
    print(f"  1. Integration Overview - Enhanced metrics with resolution rate ({len(summary_df)} integrations)")
    print(f"  2. Count by Integration - Case type & priority breakdown per IA ({len(integration_count_df)} integrations)")
    print(f"  3. Resolution Breakdown by IA - Complete resolution counts per integration ({len(resolution_breakdown_df)} integrations)")
    print(f"  4. Count by Month - Monthly trends and top integrations ({len(monthly_count_df)} months)")
    print(f"  5. Customer Analysis - Customer-specific cases with tier ({len(customer_analysis_df)} customers)")
    print(f"  6. Error Categories - Errors grouped by category ({len(error_category_summary_df)} categories)")
    print(f"  7. Error Distribution by IA - Detailed breakdown per integration ({len(error_distribution_df)} rows)")
    print(f"  8. Frequent Flow Issues - Flows mentioned 2+ times with details ({len(frequent_flows)} flows)")
    print(f"  9. Recurring Errors - Errors occurring 2+ times ({len(frequent_errors)} errors)")
    print(f" 10. Case Details - Complete extracted data ({len(cases_df)} cases)")
    
    print("\n" + "="*100)
    print("✅ ACTIONABLE DEEP DIVE COMPLETE!")
    print("="*100)
    print("\nNOTE: This is the ONLY deep dive file you need.")
    print("      Combined with 2025_Comprehensive_Analysis.xlsx, you have complete coverage:")
    print("      • Comprehensive Analysis = High-level patterns & executive summary")
    print("      • Actionable Deep Dive = Specific flows, errors, and technical details")
    print("="*100)
    
    return output_file

# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description='Deep Dive Detailed Analysis - Extract flows, mappings, and errors',
        epilog='''
Examples:
  python3 deep_dive_detailed_analysis.py --file 2025.csv
  python3 deep_dive_detailed_analysis.py --file /path/to/cases.csv --output Details.xlsx
  
This tool extracts:
  - Specific flow names from descriptions/comments
  - Field and mapping issues
  - Exact error messages
  - PRE/PRD references
  - Record types involved
        '''
    )
    
    parser.add_argument('--file', '-f', required=True, help='Path to CSV file')
    parser.add_argument('--output', '-o', help='Output Excel filename (optional)')
    
    args = parser.parse_args()
    
    try:
        deep_dive_analysis(args.file, args.output)
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()

