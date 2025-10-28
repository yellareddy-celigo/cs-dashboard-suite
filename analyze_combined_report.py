#!/usr/bin/env python3
"""
COMBINED REPORT ANALYZER
Generates a single comprehensive Excel file with all case types
Perfect for analyzing mixed case type CSV files like 2025.csv

Usage:
    python3 analyze_combined_report.py --file 2025.csv --output 2025_Comprehensive_Analysis.xlsx
"""

import pandas as pd
import argparse
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter
import re

VERSION = "1.1.0"
LAST_UPDATED = "2025-10-09"

# ============================================================================
# HELPER FUNCTIONS (Inlined from analyze_all_cases_combined.py)
# ============================================================================

def format_sheet(ws, title):
    """Apply formatting to worksheet"""
    # Header formatting
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Apply to first row if it exists
    if ws.max_row > 0:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def extract_business_use_case(description, summary):
    """Extract business use case from description or summary"""
    if pd.isna(description):
        description = ''
    if pd.isna(summary):
        summary = ''
    
    # Combine text
    full_text = str(description) + ' ' + str(summary)
    
    # Look for business use case
    if 'business use case' in full_text.lower():
        match = re.search(r'business use case[:\s]*([^h2]*)', full_text, re.IGNORECASE)
        if match:
            return match.group(1).strip()[:800]
    
    # Fallback to description
    return str(description)[:800]

def extract_customer_impact(description, summary):
    """Extract customer impact"""
    if pd.isna(description):
        description = ''
    if pd.isna(summary):
        summary = ''
    
    full_text = (str(description) + ' ' + str(summary)).lower()
    
    # Look for impact indicators
    if any(word in full_text for word in ['critical', 'urgent', 'blocking', 'stopped']):
        return 'High'
    elif any(word in full_text for word in ['important', 'affecting', 'delayed']):
        return 'Medium'
    return 'Low'

def extract_linked_items(description, summary):
    """Extract linked PRE/PRD/IO items"""
    if pd.isna(description):
        description = ''
    if pd.isna(summary):
        summary = ''
    
    full_text = str(description) + ' ' + str(summary)
    pattern = r'(PRE-\d+|PRD-\d+|IO-\d+|FEATURE-\d+)'
    matches = re.findall(pattern, full_text, re.IGNORECASE)
    return ', '.join(matches[:5]) if matches else 'None'

def extract_resolution_info(description, summary):
    """Extract resolution information"""
    if pd.isna(description):
        description = ''
    if pd.isna(summary):
        summary = ''
    
    full_text = str(description) + ' ' + str(summary)
    
    # Look for resolution info
    if 'resolution' in full_text.lower():
        match = re.search(r'resolution[:\s]*([^h2]*)', full_text, re.IGNORECASE)
        if match:
            return match.group(1).strip()[:800]
    
    return str(description)[:800]

def categorize_bug_pattern(summary, description):
    """Categorize bug patterns"""
    full_text = (str(summary) + ' ' + str(description)).lower()
    
    if any(word in full_text for word in ['auth', 'token', 'credential', 'unauthorized']):
        return 'Authentication'
    elif any(word in full_text for word in ['mapping', 'field', 'sync']):
        return 'Data Mapping/Sync'
    elif any(word in full_text for word in ['error', 'exception', 'failed']):
        return 'Code Error'
    elif any(word in full_text for word in ['config', 'setup', 'configuration']):
        return 'Configuration'
    return 'Other'

def categorize_query_pattern(summary, description):
    """Categorize query patterns"""
    full_text = (str(summary) + ' ' + str(description)).lower()
    
    if any(word in full_text for word in ['how', 'documentation', 'guide']):
        return 'How-To Question'
    elif any(word in full_text for word in ['why', 'explain', 'reason']):
        return 'Why Question'
    elif any(word in full_text for word in ['setup', 'configure', 'install']):
        return 'Configuration Question'
    return 'General Query'

def categorize_doc_enhancement_pattern(summary, description):
    """Categorize documentation/enhancement patterns"""
    full_text = (str(summary) + ' ' + str(description)).lower()
    
    if any(word in full_text for word in ['documentation', 'doc', 'guide', 'tutorial']):
        return 'Documentation Update'
    elif any(word in full_text for word in ['feature', 'enhancement', 'improve']):
        return 'Feature Enhancement'
    elif any(word in full_text for word in ['compliance', 'audit', 'report']):
        return 'Compliance/Reporting'
    return 'Other'

def extract_all_refs(row, columns_to_check):
    """Extract all PRE/PRD/IO references from specified columns"""
    all_refs = []
    source_info = []
    
    pre_prd_pattern = r'(PRE-\d+|PRD-\d+|IO-\d+|FEATURE-\d+)'
    
    for col in columns_to_check:
        if col in row.index and pd.notna(row[col]):
            text = str(row[col])
            matches = re.findall(pre_prd_pattern, text, re.IGNORECASE)
            if matches:
                unique_matches = set([m.upper() for m in matches])
                all_refs.extend(unique_matches)
                source_info.append(f"{col.split('(')[0].strip()} ({len(unique_matches)})")
    
    return list(set(all_refs)), source_info

def analyze_code_fixes_with_links(df):
    """Analyze 'Done' cases to identify code fixes with linked PRE/PRD/IO tickets"""
    
    # Filter for "Done" resolution
    done_cases = df[df['Resolution'] == 'Done'].copy()
    
    if len(done_cases) == 0:
        return None
    
    # Columns to check for linked items
    link_columns = [
        'Inward issue link (Resolves)',
        'Outward issue link (Relates)',
        'Inward issue link (Relates)',
        'Inward issue link (Problem/Incident)',
        'Inward issue link (Dependencies)',
        'Comment',
        'Description',
        'Resolution Comment'
    ]
    
    detailed_analysis = []
    
    for idx, row in done_cases.iterrows():
        case_key = row['Issue key']
        summary = row['Summary']
        priority = row['Priority'] if pd.notna(row['Priority']) else 'N/A'
        
        # Get case type and integration
        case_type_col = [col for col in df.columns if 'case type' in col.lower() and 'custom field' in col.lower()]
        integration_col = [col for col in df.columns if 'integration app' in col.lower() and 'custom field' in col.lower()]
        
        case_type = row[case_type_col[0]] if case_type_col and pd.notna(row[case_type_col[0]]) else 'N/A'
        integration = row[integration_col[0]] if integration_col and pd.notna(row[integration_col[0]]) else 'N/A'
        
        # Extract all linked items
        all_refs, source_cols = extract_all_refs(row, link_columns)
        
        # Main linked item from "Resolves" column
        resolves_link = row['Inward issue link (Resolves)'] if 'Inward issue link (Resolves)' in row.index and pd.notna(row['Inward issue link (Resolves)']) else 'None'
        
        # Check for code fix indicators
        description = str(row['Description']) if pd.notna(row['Description']) else ''
        comments = str(row['Comment']) if 'Comment' in row.index and pd.notna(row['Comment']) else ''
        resolution_comment = str(row['Resolution Comment']) if 'Resolution Comment' in row.index and pd.notna(row['Resolution Comment']) else ''
        
        all_text = f"{description} {comments} {resolution_comment}".upper()
        
        # Classification logic
        if resolves_link != 'None' and ('PRE-' in resolves_link or 'PRD-' in resolves_link or 'IO-' in resolves_link):
            classification = '✅ Code Fix (Resolves Link)'
            confidence = 'High'
            action = 'Document fix details'
        elif len(all_refs) > 0:
            classification = '✅ Code Fix (Related Links)'
            confidence = 'High'
            action = 'Document fix details'
        elif 'DEPLOY' in all_text or 'HOTFIX' in all_text:
            classification = '🟡 Likely Code Fix (Deploy/Hotfix)'
            confidence = 'Medium'
            action = 'Verify and document'
        elif 'BUG FIX' in all_text or 'FIXED IN' in all_text:
            classification = '🟡 Likely Code Fix (Fix Mentioned)'
            confidence = 'Medium'
            action = 'Verify and document'
        else:
            classification = '❌ Not Code Fix / Unknown'
            confidence = 'Low'
            action = 'Review or skip'
        
        detailed_analysis.append({
            'Case Key': case_key,
            'Classification': classification,
            'Confidence': confidence,
            'Resolves Link': resolves_link,
            'All Linked Items': ', '.join(all_refs) if all_refs else 'None',
            'Link Count': len(all_refs),
            'Link Sources': ', '.join(source_cols) if source_cols else 'None',
            'Case Type': case_type,
            'Priority': priority,
            'Integration': integration,
            'Summary': summary,
            'Action Required': action
        })
    
    return pd.DataFrame(detailed_analysis)

def analyze_combined_report(csv_file, output_file=None):
    """Analyze all case types and generate a single comprehensive report."""
    
    print("="*100)
    print(f"COMBINED COMPREHENSIVE REPORT ANALYZER")
    print(f"Version: {VERSION} | Last Updated: {LAST_UPDATED}")
    print("="*100)
    
    # Load CSV
    df = pd.read_csv(csv_file)
    print(f"\n✅ Loaded {len(df)} total cases from {csv_file}")
    print(f"Columns: {len(df.columns)}")
    
    # Identify columns
    key_col = 'Issue key'
    summary_col = 'Summary'
    priority_col = 'Priority'
    status_col = 'Status'
    created_col = 'Created'
    resolved_col = 'Resolved'
    resolution_col = 'Resolution'
    assignee_col = 'Assignee'
    description_col = 'Description'
    
    company_cols = [col for col in df.columns if 'customer' in col.lower() and 'custom field' in col.lower() and 'old' not in col.lower()]
    integration_cols = [col for col in df.columns if 'integration app' in col.lower() and 'custom field' in col.lower()]
    case_type_cols = [col for col in df.columns if 'case type' in col.lower() and 'custom field' in col.lower()]
    
    company_col = company_cols[0] if company_cols else None
    integration_col = integration_cols[0] if integration_cols else None
    case_type_col = case_type_cols[0] if case_type_cols else None
    
    if not case_type_col:
        print("\n❌ Error: No 'Case Type' column found in CSV")
        print("This script requires a Case Type column to categorize cases")
        sys.exit(1)
    
    print(f"\nIdentified columns:")
    print(f"  Company: {company_col}")
    print(f"  Integration: {integration_col}")
    print(f"  Case Type: {case_type_col}")
    
    # Get case type distribution
    case_type_dist = df[case_type_col].value_counts().to_dict()
    print(f"\n📊 Case Type Distribution:")
    for ct, count in case_type_dist.items():
        print(f"  {ct}: {count} cases")
    
    # Process all cases
    cases_data = []
    pattern_counter = Counter()
    bug_pattern_counter = Counter()
    query_pattern_counter = Counter()
    doc_pe_pattern_counter = Counter()
    
    for idx, row in df.iterrows():
        case_key = row[key_col]
        summary = row[summary_col]
        priority = row[priority_col] if pd.notna(row[priority_col]) else 'P3'
        status = row[status_col]
        created = row[created_col]
        resolved = row[resolved_col] if pd.notna(row[resolved_col]) else 'N/A'
        resolution = row[resolution_col] if pd.notna(row[resolution_col]) else 'N/A'
        assignee = row[assignee_col] if pd.notna(row[assignee_col]) else 'Unassigned'
        company = row[company_col] if company_col and pd.notna(row[company_col]) else 'N/A'
        integration = row[integration_col] if integration_col and pd.notna(row[integration_col]) else 'N/A'
        description = row[description_col] if pd.notna(row[description_col]) else ''
        case_type = row[case_type_col] if pd.notna(row[case_type_col]) else 'Unknown'
        
        business_use_case = extract_business_use_case(description, summary)
        customer_impact = extract_customer_impact(description, summary)
        how_addressed = extract_resolution_info(description, summary)
        linked_items = extract_linked_items(description, summary)
        
        # Categorize based on case type
        if case_type.lower() == 'bug':
            pattern = categorize_bug_pattern(summary, description)
            bug_pattern_counter[pattern] += 1
        elif case_type.lower() == 'query':
            pattern = categorize_query_pattern(summary, description)
            query_pattern_counter[pattern] += 1
        elif case_type.lower() in ['documentation', 'product enhancement']:
            pattern = categorize_doc_enhancement_pattern(summary, description)
            doc_pe_pattern_counter[pattern] += 1
        else:
            pattern = "Other"
        
        pattern_counter[pattern] += 1
        
        cases_data.append({
            'Case Key': case_key,
            'Case Type': case_type,
            'Summary': summary,
            'Priority': priority,
            'Status': status,
            'Created': created,
            'Resolved': resolved,
            'Resolution': resolution,
            'Company': company,
            'Integration': integration,
            'Business Use Case': business_use_case,
            'Customer Impact': customer_impact,
            'Pattern': pattern,
            'How Addressed': how_addressed,
            'Linked Work Items': linked_items,
            'Assignee': assignee
        })
    
    cases_df = pd.DataFrame(cases_data)
    
    # ============================================================================
    # OVERALL EXECUTIVE SUMMARY
    # ============================================================================
    
    total_cases = len(cases_df)
    priority_dist = cases_df['Priority'].value_counts().to_dict()
    case_type_distribution = cases_df['Case Type'].value_counts().to_dict()
    
    open_statuses = ['Open', 'On hold', 'Waiting for CS/Customer inputs', 'In Progress', 'Pending Investigation', 'Under Investigation']
    open_count = cases_df[cases_df['Status'].isin(open_statuses)].shape[0]
    closed_count = total_cases - open_count
    
    unique_companies = cases_df[cases_df['Company'].notna() & (cases_df['Company'] != 'N/A')]['Company'].nunique()
    unique_integrations = cases_df[cases_df['Integration'].notna() & (cases_df['Integration'] != 'N/A')]['Integration'].nunique()
    
    # Get resolution breakdown from original dataframe - show ALL 10 types
    resolution_col = 'Resolution'
    resolution_counts = df[resolution_col].value_counts()
    all_resolutions = resolution_counts.head(10)  # Get all 10 resolution types
    
    # Extract filename for display
    import os
    csv_filename = os.path.basename(csv_file).replace('.csv', '')
    
    overall_summary = [
        # Resolution Breakdown (ALL 10 types) - MOVED TO TOP
        {'Metric': '✅ RESOLUTION BREAKDOWN', 'Value': '', 'Percentage': '', 'Details': ''},
        {'Metric': 'Total Cases', 'Value': total_cases, 'Percentage': '100%', 'Details': f'All cases from {csv_filename}'},
    ]
    
    # Add all resolution types dynamically
    for idx, (resolution, count) in enumerate(all_resolutions.items()):
        # Use tree characters for all but the last one
        if idx < len(all_resolutions) - 1:
            prefix = '├─'
        else:
            prefix = '└─'
        
        overall_summary.append({
            'Metric': f'{prefix} {resolution}',
            'Value': count,
            'Percentage': f"{(count/total_cases*100):.1f}%",
            'Details': f'Rank #{idx+1}'
        })
    
    overall_summary.extend([
        {'Metric': '', 'Value': '', 'Percentage': '', 'Details': ''},
        
        # Status Breakdown
        {'Metric': '📈 STATUS BREAKDOWN', 'Value': '', 'Percentage': '', 'Details': ''},
        {'Metric': 'Open Cases', 'Value': open_count, 'Percentage': f"{(open_count/total_cases*100):.1f}%", 'Details': 'Cases requiring attention'},
        {'Metric': 'Closed Cases', 'Value': closed_count, 'Percentage': f"{(closed_count/total_cases*100):.1f}%", 'Details': 'Cases resolved'},
        {'Metric': 'Resolution Rate', 'Value': f"{(closed_count/total_cases*100):.1f}%", 'Percentage': '', 'Details': f"{closed_count} of {total_cases} cases closed"},
        {'Metric': '', 'Value': '', 'Percentage': '', 'Details': ''},
        
        # Priority Distribution
        {'Metric': '🎯 PRIORITY DISTRIBUTION', 'Value': '', 'Percentage': '', 'Details': ''},
        {'Metric': 'P1 (Critical/Urgent)', 'Value': priority_dist.get('P1', 0), 'Percentage': f"{(priority_dist.get('P1', 0)/total_cases*100):.1f}%", 'Details': 'Highest priority - immediate action'},
        {'Metric': 'P2 (High)', 'Value': priority_dist.get('P2', 0), 'Percentage': f"{(priority_dist.get('P2', 0)/total_cases*100):.1f}%", 'Details': 'High priority - near-term action'},
        {'Metric': 'P3 (Medium)', 'Value': priority_dist.get('P3', 0), 'Percentage': f"{(priority_dist.get('P3', 0)/total_cases*100):.1f}%", 'Details': 'Medium priority - scheduled action'},
        {'Metric': 'P4 (Low)', 'Value': priority_dist.get('P4', 0), 'Percentage': f"{(priority_dist.get('P4', 0)/total_cases*100):.1f}%", 'Details': 'Low priority - as time permits'}
    ])
    
    # Will add Case Type Breakdown after calculating bug/query/doc data below
    # (Placeholder - will be filled after case type calculations)
    
    # ============================================================================
    # CASE TYPE SUMMARIES - COMBINED
    # ============================================================================
    
    # Bug Summary
    bug_cases = cases_df[cases_df['Case Type'] == 'Bug']
    bug_open = bug_cases[bug_cases['Status'].isin(open_statuses)].shape[0]
    bug_closed = len(bug_cases) - bug_open
    bug_pattern_counter = Counter(bug_cases['Pattern'].dropna())
    
    # Query Summary
    query_cases = cases_df[cases_df['Case Type'] == 'Query']
    query_open = query_cases[query_cases['Status'].isin(open_statuses)].shape[0]
    query_closed = len(query_cases) - query_open
    query_pattern_counter = Counter(query_cases['Pattern'].dropna())
    
    # Doc/PE Summary
    doc_pe_cases = cases_df[cases_df['Case Type'].isin(['Documentation', 'Product Enhancement'])]
    doc_pe_open = doc_pe_cases[doc_pe_cases['Status'].isin(open_statuses)].shape[0]
    doc_pe_closed = len(doc_pe_cases) - doc_pe_open
    doc_pe_pattern_counter = Counter(doc_pe_cases['Pattern'].dropna())
    
    # Now add Case Type Breakdown to Overall Summary (merged as 4th section)
    overall_summary.extend([
        {'Metric': '', 'Value': '', 'Percentage': '', 'Details': ''},
        {'Metric': '📊 CASE TYPE BREAKDOWN', 'Value': '', 'Percentage': '', 'Details': ''},
        # Bug Section
        {'Metric': '🐛 Bug Cases', 'Value': len(bug_cases), 'Percentage': f"{(len(bug_cases)/total_cases*100):.1f}%", 'Details': f"{bug_open} open, {bug_closed} closed"},
        {'Metric': '❓ Query Cases', 'Value': len(query_cases), 'Percentage': f"{(len(query_cases)/total_cases*100):.1f}%", 'Details': f"{query_open} open, {query_closed} closed"},
        {'Metric': '📚 Doc/Enhancement', 'Value': len(doc_pe_cases), 'Percentage': f"{(len(doc_pe_cases)/total_cases*100):.1f}%", 'Details': f"{doc_pe_open} open, {doc_pe_closed} closed"}
    ])
    
    overall_df = pd.DataFrame(overall_summary)
    
    # ============================================================================
    # PATTERN ANALYSIS
    # ============================================================================
    
    pattern_analysis = []
    for pattern, count in pattern_counter.most_common():
        pattern_cases = cases_df[cases_df['Pattern'] == pattern]
        case_keys = ', '.join(pattern_cases['Case Key'].tolist()[:5])
        case_types = pattern_cases['Case Type'].value_counts().to_dict()
        case_type_str = ', '.join([f"{ct}: {c}" for ct, c in case_types.items()])
        
        pattern_priorities = pattern_cases['Priority'].value_counts()
        priority_str = ', '.join([f"{p}: {c}" for p, c in pattern_priorities.items()])
        
        pattern_analysis.append({
            'Pattern': pattern,
            'Case Count': count,
            'Percentage': f"{(count/total_cases)*100:.1f}%",
            'Case Types': case_type_str,
            'Sample Cases': case_keys,
            'Priority Distribution': priority_str
        })
    
    pattern_df = pd.DataFrame(pattern_analysis)
    
    # ============================================================================
    # TOP INTEGRATIONS
    # ============================================================================
    
    integration_analysis = []
    integration_dist = cases_df[cases_df['Integration'] != 'N/A']['Integration'].value_counts().head(20)
    
    for integration, count in integration_dist.items():
        int_cases = cases_df[cases_df['Integration'] == integration]
        int_open = int_cases[int_cases['Status'].isin(open_statuses)].shape[0]
        case_types = int_cases['Case Type'].value_counts().to_dict()
        case_type_str = ', '.join([f"{ct}: {c}" for ct, c in case_types.items()])
        
        integration_analysis.append({
            'Integration': integration,
            'Total Cases': count,
            'Open Cases': int_open,
            'Closed Cases': count - int_open,
            'Case Types': case_type_str,
            'Top Pattern': int_cases['Pattern'].value_counts().index[0] if len(int_cases) > 0 else 'N/A'
        })
    
    integration_df = pd.DataFrame(integration_analysis)
    
    # ============================================================================
    # CREATE EXCEL FILE
    # ============================================================================
    
    print(f"\n📊 Analysis Summary:")
    print(f"  Total Cases: {total_cases}")
    print(f"  Bug Cases: {len(bug_cases)}")
    print(f"  Query Cases: {len(query_cases)}")
    print(f"  Doc/PE Cases: {len(doc_pe_cases)}")
    print(f"  Open Cases: {open_count} ({(open_count/total_cases*100):.1f}%)")
    print(f"  Patterns Identified: {len(pattern_counter)}")
    print(f"  Unique Integrations: {unique_integrations}")
    
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'Comprehensive_Analysis_{timestamp}.xlsx'
    
    print(f"\n📝 Creating comprehensive Excel file: {output_file}")
    
    # Analyze code fixes with linked items
    code_fix_df = analyze_code_fixes_with_links(df)
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write all sheets
        overall_df.to_excel(writer, sheet_name='Overall Summary', index=False, startrow=1)
        cases_df.to_excel(writer, sheet_name='All Cases', index=False, startrow=1)
        pattern_df.to_excel(writer, sheet_name='Pattern Analysis', index=False, startrow=1)
        integration_df.to_excel(writer, sheet_name='Top Integrations', index=False, startrow=1)
        
        # Format all sheets
        workbook = writer.book
        format_sheet(workbook['Overall Summary'], "Overall Summary - Complete Analysis")
        format_sheet(workbook['All Cases'], "All Cases - Complete Details")
        format_sheet(workbook['Pattern Analysis'], "Pattern Analysis - All Case Types")
        format_sheet(workbook['Top Integrations'], "Top 20 Integrations by Case Count")
        
        # Apply special formatting to Overall Summary section headers
        ws_overall = workbook['Overall Summary']
        section_header_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        section_header_font = Font(bold=True, size=11, color="1F4E78")
        spacer_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        for row in ws_overall.iter_rows(min_row=3, max_row=ws_overall.max_row):
            metric_value = str(row[0].value)
            # Format section headers (rows with emojis in Metric column)
            if '📈' in metric_value or '🎯' in metric_value or '✅' in metric_value or '📊' in metric_value or '🐛' in metric_value or '❓' in metric_value or '📚' in metric_value:
                for cell in row:
                    cell.fill = section_header_fill
                    cell.font = section_header_font
            # Format spacer rows (empty rows)
            elif metric_value == '' or metric_value == 'None':
                for cell in row:
                    cell.fill = spacer_fill
        
        # Highlight "All cases from 20XX" cell
        for row in ws_overall.iter_rows(min_row=3, max_row=ws_overall.max_row):
            for cell in row:
                if cell.value and ('All cases from' in str(cell.value)):
                    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    highlight_font = Font(bold=True, size=11, color='000000')
                    cell.fill = highlight_fill
                    cell.font = highlight_font
                    break
    
    # Add "Code Fix with Links" sheet if there are Done cases
    if code_fix_df is not None:
        from openpyxl import load_workbook
        
        book = load_workbook(output_file)
        book.create_sheet('Code Fix with Links', 5)
        ws = book['Code Fix with Links']
        
        # Add title - LEFT ALIGNED
        ws['A1'] = 'Done Cases - Code Fix Analysis with Linked PRE/PRD/IO Tickets'
        ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
        ws.merge_cells('A1:L1')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Write headers
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for idx, header in enumerate(code_fix_df.columns, 1):
            cell = ws.cell(row=2, column=idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write data
        for r_idx, row in code_fix_df.iterrows():
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx + 3, column=c_idx)
                cell.value = value
        
        # Apply alignment
        for row_idx in range(3, ws.max_row + 1):
            for col_idx in range(1, 13):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx in [1, 2, 3, 6, 8, 9]:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Apply color coding
        code_fix_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        likely_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        not_fix_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        for row_idx in range(3, ws.max_row + 1):
            classification = str(ws.cell(row=row_idx, column=2).value)
            
            if '✅' in classification:
                for col_idx in range(1, 13):
                    ws.cell(row=row_idx, column=col_idx).fill = code_fix_fill
            elif '🟡' in classification:
                for col_idx in range(1, 13):
                    ws.cell(row=row_idx, column=col_idx).fill = likely_fill
            elif '❌' in classification:
                for col_idx in range(1, 13):
                    ws.cell(row=row_idx, column=col_idx).fill = not_fix_fill
        
        # Set column widths
        column_widths = {1: 15, 2: 30, 3: 12, 4: 18, 5: 35, 6: 12, 7: 40, 8: 15, 9: 10, 10: 30, 11: 60, 12: 25}
        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 35
        
        book.save(output_file)
        
        # Print stats
        code_fix_cases = code_fix_df[code_fix_df['Classification'].str.contains('✅')]
        print(f"\n✅ Comprehensive Excel file created successfully!")
        print(f"\n📋 File contains 5 sheets:")
        print(f"  1. Overall Summary - 4 sections (Resolution, Status, Priority, Case Type Breakdown)")
        print(f"  2. All Cases - All {total_cases} cases with 16 columns")
        print(f"  3. Pattern Analysis - {len(pattern_counter)} patterns identified")
        print(f"  4. Top Integrations - Top 20 integrations analysis")
        print(f"  5. Code Fix with Links - {len(code_fix_df)} Done cases ({len(code_fix_cases)} confirmed code fixes)")
    else:
        print(f"\n✅ Comprehensive Excel file created successfully!")
        print(f"\n📋 File contains 4 sheets:")
        print(f"  1. Overall Summary - 4 sections (Resolution, Status, Priority, Case Type Breakdown)")
        print(f"  2. All Cases - All {total_cases} cases with 16 columns")
        print(f"  3. Pattern Analysis - {len(pattern_counter)} patterns identified")
        print(f"  4. Top Integrations - Top 20 integrations analysis")
        print(f"\n⚠️  No 'Done' cases found - Code Fix with Links sheet not created")
    
    print("\n" + "="*100)
    print("COMPREHENSIVE ANALYSIS COMPLETE!")
    print("="*100)
    
    return output_file

# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description='Combined Report Analyzer - Single comprehensive file for all case types',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  python3 analyze_combined_report.py --file 2025.csv
  python3 analyze_combined_report.py --file /path/to/cases.csv --output MyReport.xlsx
  
This tool generates a single Excel file with:
  - Overall summary across all case types
  - Combined case type breakdown (Bug, Query, Doc/PE in one sheet)
  - Complete case details (all cases in one sheet)
  - Pattern analysis
  - Top integrations analysis
  - Code fixes with linked PRE/PRD/IO tickets (if Done cases exist)
        '''
    )
    
    parser.add_argument('--file', '-f', required=True, help='Path to CSV file with cases')
    parser.add_argument('--output', '-o', help='Output Excel filename (optional, auto-generated if not provided)')
    parser.add_argument('--version', action='version', version=f'%(prog)s {VERSION}')
    
    args = parser.parse_args()
    
    try:
        analyze_combined_report(args.file, args.output)
    except FileNotFoundError:
        print(f"\n❌ Error: File not found: {args.file}")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Error during analysis: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

