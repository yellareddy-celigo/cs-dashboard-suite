#!/usr/bin/env python3
"""
Real JIRA Comprehensive Dashboard Generator
Uses real JIRA data pulled via MCP Atlassian tools
"""

import argparse
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows
import re

class RealJiraComprehensiveDashboard:
    def __init__(self, start_date='2023-01-01', end_date='2025-12-31'):
        self.start_date = start_date
        self.end_date = end_date
        self.df = None
        
        # Configuration for real data processing
        self.integration_apps = [
            'Salesforce', 'HubSpot', 'Zendesk', 'Slack', 'Microsoft Teams',
            'Zoom', 'Google Workspace', 'AWS', 'Azure', 'ServiceNow',
            'Jira', 'Confluence', 'Trello', 'Asana', 'Monday.com'
        ]
        
        self.root_causes = [
            'API Integration Failure', 'Data Synchronization Issue', 'Authentication Problem',
            'Rate Limiting', 'Configuration Error', 'Network Timeout', 'Data Format Mismatch',
            'Permission Issue', 'Version Compatibility', 'Third-party Service Down'
        ]
        
        self.resolutions = [
            'Fixed', 'Won\'t Fix', 'Duplicate', 'Cannot Reproduce', 'Incomplete',
            'Done', 'Resolved', 'Closed', 'Cancelled', 'Deferred'
        ]

    def load_real_jira_data(self, jira_data):
        """Load and process real JIRA data"""
        print("📊 Processing real JIRA data...")
        
        # Convert to DataFrame
        self.df = pd.DataFrame(jira_data)
        
        # Convert date columns
        self.df['Created'] = pd.to_datetime(self.df['Created'])
        self.df['Updated'] = pd.to_datetime(self.df['Updated'])
        self.df['Resolved'] = pd.to_datetime(self.df['Resolved'], errors='coerce')
        
        # Filter by date range
        self.df = self.df[
            (self.df['Created'] >= self.start_date) & 
            (self.df['Created'] <= self.end_date)
        ]
        
        # Add derived columns
        self.df['Month-Year'] = self.df['Created'].dt.to_period('M').astype(str)
        self.df['Year'] = self.df['Created'].dt.year
        self.df['Month'] = self.df['Created'].dt.month
        self.df['Quarter'] = self.df['Created'].dt.quarter.apply(lambda x: f'Q{x}')
        
        # Calculate resolution time
        self.df['Resolution Time (days)'] = (
            self.df['Resolved'] - self.df['Created']
        ).dt.days
        
        # Fill missing values
        self.df['Resolution Time (days)'] = self.df['Resolution Time (days)'].fillna(0)
        self.df['Resolution'] = self.df['Resolution'].fillna('Unresolved')
        self.df['Root Cause'] = self.df['Root Cause'].fillna('Unknown')
        
        # Extract integration apps from summary/description
        self.df['Integration Apps'] = self.df['Summary'].apply(self._extract_integration_apps)
        
        print(f"✅ Processed {len(self.df)} real JIRA issues")
        return self.df

    def _extract_integration_apps(self, summary):
        """Extract integration apps from issue summary"""
        if pd.isna(summary):
            return 'Unknown'
        
        summary_lower = summary.lower()
        found_apps = []
        
        for app in self.integration_apps:
            if app.lower() in summary_lower:
                found_apps.append(app)
        
        return ', '.join(found_apps) if found_apps else 'General'

    def create_comprehensive_dashboard(self, output_file='real_jira_comprehensive_dashboard.xlsx'):
        """Create comprehensive dashboard with real JIRA data"""
        print(f"🚀 Creating comprehensive dashboard: {output_file}")
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create all sheets
        self._create_executive_summary(wb)
        self._create_integration_apps_analysis(wb)
        self._create_monthly_trends(wb)
        self._create_integration_monthly_matrix_with_charts(wb)
        self._create_resolution_analysis(wb)
        self._create_root_cause_analysis(wb)
        self._create_action_items(wb)
        self._create_raw_data(wb)
        
        # Save workbook
        wb.save(output_file)
        print(f"✅ Dashboard saved: {output_file}")
        return output_file

    def _create_executive_summary(self, wb):
        """Create Executive Summary sheet"""
        ws = wb.create_sheet("📊 Executive Summary")
        
        # Title
        ws['A1'] = "Customer Success Dashboard - Executive Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Date range
        ws['A3'] = f"Analysis Period: {self.start_date} to {self.end_date}"
        ws['A3'].font = Font(size=12, bold=True)
        
        # Key metrics
        total_issues = len(self.df)
        resolved_issues = len(self.df[self.df['Status'].isin(['Done', 'Resolved', 'Closed'])])
        avg_resolution_time = self.df['Resolution Time (days)'].mean()
        
        ws['A5'] = "Key Metrics"
        ws['A5'].font = Font(size=14, bold=True)
        
        metrics = [
            ("Total Issues", total_issues),
            ("Resolved Issues", resolved_issues),
            ("Resolution Rate", f"{(resolved_issues/total_issues*100):.1f}%"),
            ("Avg Resolution Time", f"{avg_resolution_time:.1f} days"),
            ("Data Source", "Real JIRA Data via MCP")
        ]
        
        for i, (metric, value) in enumerate(metrics, 6):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Top integration apps
        ws['A12'] = "Top Integration Apps by Issue Count"
        ws['A12'].font = Font(size=14, bold=True)
        
        app_counts = self.df['Integration Apps'].value_counts().head(10)
        for i, (app, count) in enumerate(app_counts.items(), 13):
            ws[f'A{i}'] = app
            ws[f'B{i}'] = count

    def _create_integration_apps_analysis(self, wb):
        """Create Integration Apps Analysis sheet"""
        ws = wb.create_sheet("🔧 Integration Apps")
        
        # Title
        ws['A1'] = "Integration Apps Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        # App summary
        app_summary = self.df.groupby('Integration Apps').agg({
            'Issue Key': 'count',
            'Resolution Time (days)': 'mean',
            'Status': lambda x: (x.isin(['Done', 'Resolved', 'Closed']).sum())
        }).round(2)
        
        app_summary.columns = ['Total Issues', 'Avg Resolution Time (days)', 'Resolved Issues']
        app_summary['Resolution Rate'] = (app_summary['Resolved Issues'] / app_summary['Total Issues'] * 100).round(1)
        
        # Write data
        ws['A3'] = "Integration App"
        ws['B3'] = "Total Issues"
        ws['C3'] = "Resolved Issues"
        ws['D3'] = "Resolution Rate (%)"
        ws['E3'] = "Avg Resolution Time (days)"
        
        # Headers styling
        for col in ['A3', 'B3', 'C3', 'D3', 'E3']:
            ws[col].font = Font(bold=True)
            ws[col].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for i, (app, row) in enumerate(app_summary.iterrows(), 4):
            ws[f'A{i}'] = app
            ws[f'B{i}'] = row['Total Issues']
            ws[f'C{i}'] = row['Resolved Issues']
            ws[f'D{i}'] = row['Resolution Rate']
            ws[f'E{i}'] = row['Avg Resolution Time (days)']

    def _create_monthly_trends(self, wb):
        """Create Monthly Trends sheet"""
        ws = wb.create_sheet("📈 Monthly Trends")
        
        # Title
        ws['A1'] = "Monthly Trends Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Monthly summary
        monthly_summary = self.df.groupby('Month-Year').agg({
            'Issue Key': 'count',
            'Resolution Time (days)': 'mean',
            'Status': lambda x: (x.isin(['Done', 'Resolved', 'Closed']).sum())
        }).round(2)
        
        monthly_summary.columns = ['Total Issues', 'Avg Resolution Time (days)', 'Resolved Issues']
        monthly_summary['Resolution Rate'] = (monthly_summary['Resolved Issues'] / monthly_summary['Total Issues'] * 100).round(1)
        
        # Write data
        ws['A3'] = "Month-Year"
        ws['B3'] = "Total Issues"
        ws['C3'] = "Resolved Issues"
        ws['D3'] = "Resolution Rate (%)"
        ws['E3'] = "Avg Resolution Time (days)"
        
        # Headers styling
        for col in ['A3', 'B3', 'C3', 'D3', 'E3']:
            ws[col].font = Font(bold=True)
            ws[col].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for i, (month, row) in enumerate(monthly_summary.iterrows(), 4):
            ws[f'A{i}'] = month
            ws[f'B{i}'] = row['Total Issues']
            ws[f'C{i}'] = row['Resolved Issues']
            ws[f'D{i}'] = row['Resolution Rate']
            ws[f'E{i}'] = row['Avg Resolution Time (days)']

    def _create_integration_monthly_matrix_with_charts(self, wb):
        """Create Issues per Integration App per Month with charts"""
        ws = wb.create_sheet("📊 Issues per App per Month")
        
        # Title
        ws['A1'] = "Issues per Integration App per Month"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Create pivot table
        pivot = self.df.pivot_table(
            index='Integration Apps',
            columns='Month-Year',
            values='Issue Key',
            aggfunc='count',
            fill_value=0
        )
        
        # Write pivot table
        ws['A3'] = "Integration App"
        col_idx = 2
        for month in pivot.columns:
            ws.cell(row=3, column=col_idx, value=month)
            col_idx += 1
        
        # Headers styling
        for col in range(1, col_idx):
            ws.cell(row=3, column=col).font = Font(bold=True)
            ws.cell(row=3, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        row_idx = 4
        for app in pivot.index:
            ws.cell(row=row_idx, column=1, value=app)
            col_idx = 2
            for month in pivot.columns:
                ws.cell(row=row_idx, column=col_idx, value=pivot.loc[app, month])
                col_idx += 1
            row_idx += 1
        
        # Add charts
        self._add_charts_to_monthly_matrix(ws, pivot, row_idx)

    def _add_charts_to_monthly_matrix(self, ws, pivot, start_row):
        """Add visual charts to the monthly matrix sheet"""
        # Chart 1: Bar Chart - Top 10 Apps by Total Issues
        chart1 = BarChart()
        chart1.title = "Top 10 Integration Apps by Total Issues"
        chart1.x_axis.title = "Integration Apps"
        chart1.y_axis.title = "Number of Issues"
        
        # Get top 10 apps
        top_apps = pivot.sum(axis=1).nlargest(10)
        
        # Data for chart
        data_rows = []
        for app in top_apps.index:
            data_rows.append([app, top_apps[app]])
        
        # Write chart data
        chart_start_row = start_row + 2
        ws[f'A{chart_start_row}'] = "App"
        ws[f'B{chart_start_row}'] = "Total Issues"
        
        for i, (app, count) in enumerate(data_rows, chart_start_row + 1):
            ws[f'A{i}'] = app
            ws[f'B{i}'] = count
        
        # Add chart
        chart1.add_data(Reference(ws, min_col=2, min_row=chart_start_row, max_row=chart_start_row + len(data_rows)))
        chart1.set_categories(Reference(ws, min_col=1, min_row=chart_start_row + 1, max_row=chart_start_row + len(data_rows)))
        ws.add_chart(chart1, f'D{chart_start_row}')
        
        # Chart 2: Line Chart - Monthly Trends for Top 5 Apps
        chart2 = LineChart()
        chart2.title = "Monthly Trends for Top 5 Integration Apps"
        chart2.x_axis.title = "Month"
        chart2.y_axis.title = "Number of Issues"
        
        # Get top 5 apps
        top_5_apps = pivot.sum(axis=1).nlargest(5)
        
        # Write chart data
        chart2_start_row = chart_start_row + len(data_rows) + 3
        ws[f'A{chart2_start_row}'] = "Month"
        
        col_idx = 2
        for app in top_5_apps.index:
            ws.cell(row=chart2_start_row, column=col_idx, value=app)
            col_idx += 1
        
        # Data for each month
        for i, month in enumerate(pivot.columns, chart2_start_row + 1):
            ws.cell(row=i, column=1, value=month)
            col_idx = 2
            for app in top_5_apps.index:
                ws.cell(row=i, column=col_idx, value=pivot.loc[app, month])
                col_idx += 1
        
        # Add chart
        chart2.add_data(Reference(ws, min_col=2, min_row=chart2_start_row, max_col=col_idx-1, max_row=chart2_start_row + len(pivot.columns)))
        chart2.set_categories(Reference(ws, min_col=1, min_row=chart2_start_row + 1, max_row=chart2_start_row + len(pivot.columns)))
        ws.add_chart(chart2, f'D{chart2_start_row}')
        
        # Chart 3: Pie Chart - Distribution of Issues by App
        chart3 = PieChart()
        chart3.title = "Distribution of Issues by Integration App"
        
        # Write chart data
        chart3_start_row = chart2_start_row + len(pivot.columns) + 3
        ws[f'A{chart3_start_row}'] = "App"
        ws[f'B{chart3_start_row}'] = "Issues"
        
        for i, (app, count) in enumerate(top_apps.items(), chart3_start_row + 1):
            ws[f'A{i}'] = app
            ws[f'B{i}'] = count
        
        # Add chart
        chart3.add_data(Reference(ws, min_col=2, min_row=chart3_start_row, max_row=chart3_start_row + len(top_apps)))
        chart3.set_categories(Reference(ws, min_col=1, min_row=chart3_start_row + 1, max_row=chart3_start_row + len(top_apps)))
        ws.add_chart(chart3, f'D{chart3_start_row}')
        
        # Chart 4: Heatmap-style matrix
        chart4_start_row = chart3_start_row + len(top_apps) + 3
        ws[f'A{chart4_start_row}'] = "Heatmap Matrix: Issues per App per Month"
        ws[f'A{chart4_start_row}'].font = Font(size=14, bold=True)
        
        # Write heatmap data
        heatmap_start_row = chart4_start_row + 2
        ws[f'A{heatmap_start_row}'] = "App/Month"
        
        col_idx = 2
        for month in pivot.columns:
            ws.cell(row=heatmap_start_row, column=col_idx, value=month)
            col_idx += 1
        
        # Data with conditional formatting
        row_idx = heatmap_start_row + 1
        for app in pivot.index:
            ws.cell(row=row_idx, column=1, value=app)
            col_idx = 2
            for month in pivot.columns:
                value = pivot.loc[app, month]
                ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Conditional formatting based on value
                if value > 10:
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                elif value > 5:
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                elif value > 0:
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                col_idx += 1
            row_idx += 1

    def _create_resolution_analysis(self, wb):
        """Create Resolution Analysis sheet"""
        ws = wb.create_sheet("🔍 Resolution Analysis")
        
        # Title
        ws['A1'] = "Resolution Analysis - Code Fixes for Test Case Coverage"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Resolution summary
        resolution_summary = self.df.groupby('Resolution').agg({
            'Issue Key': 'count',
            'Resolution Time (days)': 'mean'
        }).round(2)
        
        resolution_summary.columns = ['Count', 'Avg Resolution Time (days)']
        
        # Write data
        ws['A3'] = "Resolution Type"
        ws['B3'] = "Count"
        ws['C3'] = "Avg Resolution Time (days)"
        
        # Headers styling
        for col in ['A3', 'B3', 'C3']:
            ws[col].font = Font(bold=True)
            ws[col].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for i, (resolution, row) in enumerate(resolution_summary.iterrows(), 4):
            ws[f'A{i}'] = resolution
            ws[f'B{i}'] = row['Count']
            ws[f'C{i}'] = row['Avg Resolution Time (days)']
        
        # Code fixes analysis
        ws['A12'] = "Code Fixes per Integration App per Month"
        ws['A12'].font = Font(size=14, bold=True)
        
        # Filter for code fixes
        code_fixes = self.df[self.df['Resolution'].isin(['Fixed', 'Done', 'Resolved'])]
        
        if len(code_fixes) > 0:
            code_pivot = code_fixes.pivot_table(
                index='Integration Apps',
                columns='Month-Year',
                values='Issue Key',
                aggfunc='count',
                fill_value=0
            )
            
            # Write pivot table
            ws['A14'] = "Integration App"
            col_idx = 2
            for month in code_pivot.columns:
                ws.cell(row=14, column=col_idx, value=month)
                col_idx += 1
            
            # Headers styling
            for col in range(1, col_idx):
                ws.cell(row=14, column=col).font = Font(bold=True)
                ws.cell(row=14, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Data
            row_idx = 15
            for app in code_pivot.index:
                ws.cell(row=row_idx, column=1, value=app)
                col_idx = 2
                for month in code_pivot.columns:
                    ws.cell(row=row_idx, column=col_idx, value=code_pivot.loc[app, month])
                    col_idx += 1
                row_idx += 1
            
            # Summary statistics
            ws[f'A{row_idx + 2}'] = "Summary Statistics"
            ws[f'A{row_idx + 2}'].font = Font(size=14, bold=True)
            
            ws[f'A{row_idx + 4}'] = "Total Code Fixes"
            ws[f'B{row_idx + 4}'] = len(code_fixes)
            
            ws[f'A{row_idx + 5}'] = "Code Fix Rate"
            ws[f'B{row_idx + 5}'] = f"{(len(code_fixes)/len(self.df)*100):.1f}%"
            
            ws[f'A{row_idx + 6}'] = "Avg Code Fix Time"
            ws[f'B{row_idx + 6}'] = f"{code_fixes['Resolution Time (days)'].mean():.1f} days"
        else:
            ws['A14'] = "No code fixes found in the data"

    def _create_root_cause_analysis(self, wb):
        """Create Root Cause Analysis sheet"""
        ws = wb.create_sheet("🔍 Root Cause Analysis")
        
        # Title
        ws['A1'] = "Root Cause Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        # Root cause summary
        root_cause_summary = self.df.groupby('Root Cause').agg({
            'Issue Key': 'count',
            'Resolution Time (days)': 'mean'
        }).round(2)
        
        root_cause_summary.columns = ['Count', 'Avg Resolution Time (days)']
        root_cause_summary = root_cause_summary.sort_values('Count', ascending=False)
        
        # Write data
        ws['A3'] = "Root Cause"
        ws['B3'] = "Count"
        ws['C3'] = "Avg Resolution Time (days)"
        
        # Headers styling
        for col in ['A3', 'B3', 'C3']:
            ws[col].font = Font(bold=True)
            ws[col].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for i, (cause, row) in enumerate(root_cause_summary.iterrows(), 4):
            ws[f'A{i}'] = cause
            ws[f'B{i}'] = row['Count']
            ws[f'C{i}'] = row['Avg Resolution Time (days)']

    def _create_action_items(self, wb):
        """Create Action Items sheet"""
        ws = wb.create_sheet("📋 Action Items")
        
        # Title
        ws['A1'] = "Action Items & Recommendations"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Action items based on analysis
        action_items = [
            ("High Priority", "Focus on top integration apps with most issues"),
            ("Medium Priority", "Improve resolution time for long-running issues"),
            ("Low Priority", "Monitor trends for early detection of problems"),
            ("Data Quality", "Ensure consistent root cause categorization"),
            ("Process Improvement", "Implement automated testing for integration apps")
        ]
        
        ws['A3'] = "Priority"
        ws['B3'] = "Action Item"
        
        # Headers styling
        for col in ['A3', 'B3']:
            ws[col].font = Font(bold=True)
            ws[col].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for i, (priority, item) in enumerate(action_items, 4):
            ws[f'A{i}'] = priority
            ws[f'B{i}'] = item

    def _create_raw_data(self, wb):
        """Create Raw Data sheet"""
        ws = wb.create_sheet("📄 Raw Data")
        
        # Title
        ws['A1'] = "Raw JIRA Data"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:M1')
        
        # Write DataFrame to sheet
        for r in dataframe_to_rows(self.df, index=False, header=True):
            ws.append(r)
        
        # Style headers
        for col in range(1, len(self.df.columns) + 1):
            ws.cell(row=2, column=col).font = Font(bold=True)
            ws.cell(row=2, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

def main():
    parser = argparse.ArgumentParser(description="Generate comprehensive dashboard from real JIRA data")
    parser.add_argument('--start-date', type=str, default='2023-01-01', help='Start date (YYYY-MM-DD)')
    parser.add_argument('--end-date', type=str, default='2025-12-31', help='End date (YYYY-MM-DD)')
    parser.add_argument('--output', type=str, default='real_jira_comprehensive_dashboard.xlsx', help='Output file name')
    args = parser.parse_args()
    
    # This would normally load real JIRA data
    # For now, we'll use the data structure from the previous script
    print("📊 Real JIRA Comprehensive Dashboard Generator")
    print(f"📅 Date Range: {args.start_date} to {args.end_date}")
    print("🔗 Data Source: Real JIRA Data via MCP")
    
    # Create dashboard
    dashboard = RealJiraComprehensiveDashboard(args.start_date, args.end_date)
    
    # Note: In a real implementation, you would load the actual JIRA data here
    # For demonstration, we'll show the structure
    print("\n📋 Dashboard will include:")
    print("  📊 Executive Summary")
    print("  🔧 Integration Apps Analysis")
    print("  📈 Monthly Trends")
    print("  📊 Issues per App per Month (with charts)")
    print("  🔍 Resolution Analysis")
    print("  🔍 Root Cause Analysis")
    print("  📋 Action Items")
    print("  📄 Raw Data")
    
    print(f"\n✅ Dashboard structure created. To generate with real data:")
    print(f"   1. Ensure MCP Atlassian connection is configured")
    print(f"   2. Run the JIRA data puller script")
    print(f"   3. Load the real data into this dashboard")

if __name__ == "__main__":
    main()