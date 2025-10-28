#!/usr/bin/env python3
"""
MASTER COMPREHENSIVE REPORT GENERATOR
Combines Comprehensive Analysis + Deep Dive Analysis into a single Excel file
Perfect for complete case analysis with both high-level and detailed insights

Usage:
    python3 generate_master_report.py --file 2025.csv --output 2025_Master_Report.xlsx
"""

import pandas as pd
import argparse
import sys
from datetime import datetime
import subprocess
import re

VERSION = "1.1.0"
LAST_UPDATED = "2025-10-09"

def extract_customer_from_description(description, summary):
    """Extract customer name from description using various patterns"""
    if pd.isna(description) and pd.isna(summary):
        return 'Unknown'
    
    full_text = str(description) + ' ' + str(summary)
    
    # Customer patterns to search for
    customer_patterns = [
        r'Company:\s*([^\n]+)',
        r'Customer:\s*([^\n]+)',
        r'Account:\s*([^\n]+)',
        r'User:\s*([^\n]+)',
        r'Client:\s*([^\n]+)',
        r'Organization:\s*([^\n]+)',
        r'Business:\s*([^\n]+)',
        r'Enterprise:\s*([^\n]+)',
        r'Customer Name:\s*([^\n]+)',
        r'Account Name:\s*([^\n]+)',
        r'Company Name:\s*([^\n]+)'
    ]
    
    for pattern in customer_patterns:
        match = re.search(pattern, full_text, re.IGNORECASE)
        if match:
            customer = match.group(1).strip()
            # Clean up the customer name
            customer = re.sub(r'\|\|.*$', '', customer)  # Remove everything after ||
            customer = re.sub(r'\(Tier \d+\)', '', customer)  # Remove tier info
            customer = customer.strip()
            # Filter out common non-customer values
            if (len(customer) > 2 and 
                customer.lower() not in ['none', 'unknown', 'n/a', 'na', 'tbd', 'to be determined', 
                                       'internal', 'test', 'demo', 'sample', 'example'] and
                not customer.startswith('h1.') and
                not customer.startswith('h2.') and
                not customer.startswith('*') and
                not customer.startswith('#')):
                return customer
    
    return 'Unknown'

def generate_master_report(csv_file, output_file=None):
    """Generate a single master report combining both analyses."""
    
    print("="*100)
    print(f"MASTER COMPREHENSIVE REPORT GENERATOR")
    print(f"Version: {VERSION} | Last Updated: {LAST_UPDATED}")
    print("="*100)
    
    if output_file is None:
        # Extract year from filename or use current year
        import re
        year_match = re.search(r'(20\d{2})', csv_file)
        year = year_match.group(1) if year_match else datetime.now().strftime('%Y')
        output_file = f'{year}_Master_Report.xlsx'
    
    # Generate temporary files for both reports
    temp_comprehensive = 'temp_comprehensive.xlsx'
    temp_deep_dive = 'temp_deep_dive.xlsx'
    
    print(f"\n📊 Step 1/3: Generating Comprehensive Analysis...")
    try:
        result = subprocess.run([
            'python3', 'analyze_combined_report.py',
            '--file', csv_file,
            '--output', temp_comprehensive
        ], capture_output=True, text=True)
        
        if result.returncode != 0:
            print(f"❌ Error in Comprehensive Analysis:")
            print(result.stderr)
            return False
        
        print("✅ Comprehensive Analysis complete")
    except Exception as e:
        print(f"❌ Error running Comprehensive Analysis: {str(e)}")
        return False
    
    print(f"\n📊 Step 2/3: Generating Deep Dive Analysis...")
    try:
        result = subprocess.run([
            'python3', 'deep_dive_detailed_analysis.py',
            '--file', csv_file,
            '--output', temp_deep_dive
        ], capture_output=True, text=True)
        
        if result.returncode != 0:
            print(f"❌ Error in Deep Dive Analysis:")
            print(result.stderr)
            return False
        
        print("✅ Deep Dive Analysis complete")
    except Exception as e:
        print(f"❌ Error running Deep Dive Analysis: {str(e)}")
        return False
    
    print(f"\n📊 Step 3/3: Adding Customer Extraction Analysis...")
    
    # Load the original CSV to extract customer information
    print(f"  Extracting customer information from descriptions...")
    df_original = pd.read_csv(csv_file)
    df_original['Extracted_Customer'] = df_original.apply(lambda row: extract_customer_from_description(
        row.get('Description', ''), row.get('Summary', '')), axis=1)
    
    # Extract resolution comments if available
    if 'Custom field (Resolution Comments)' in df_original.columns:
        df_original['Resolution_Comments'] = df_original['Custom field (Resolution Comments)']
    else:
        df_original['Resolution_Comments'] = ''
    
    # Create comprehensive customer analysis data
    customer_stats = df_original['Extracted_Customer'].value_counts()
    total_cases = len(df_original)
    cases_with_customer = len(df_original[df_original['Extracted_Customer'] != 'Unknown'])
    cases_without_customer = len(df_original[df_original['Extracted_Customer'] == 'Unknown'])

    # Create comprehensive customer analysis DataFrame
    customer_analysis_data = []
    
    # Summary statistics section
    customer_analysis_data.append(['COMPREHENSIVE CUSTOMER ANALYSIS WITH EXTRACTED DETAILS', ''])
    customer_analysis_data.append(['Metric', 'Value'])
    customer_analysis_data.append(['Total Cases Analyzed', str(total_cases)])
    customer_analysis_data.append(['Cases with Customer Info', str(cases_with_customer)])
    customer_analysis_data.append(['Cases without Customer Info', str(cases_without_customer)])
    customer_analysis_data.append(['Customer Info Coverage', f"{(cases_with_customer/total_cases*100):.1f}%"])
    customer_analysis_data.append(['Unique Customers Identified', str(len(customer_stats))])
    customer_analysis_data.append(['', ''])  # Empty row
    
    # Customer breakdown by tier and priority
    customer_analysis_data.append(['CUSTOMER BREAKDOWN BY TIER AND PRIORITY', ''])
    customer_analysis_data.append(['Customer Name', 'Tier', 'Total Cases', 'P1', 'P2', 'P3', 'P4', 'Open Cases', 'Closed Cases', 'Top Resolution', 'Top Integration', 'Case Examples'])
    
    for customer, count in customer_stats.head(30).items():
        if customer != 'Unknown' and pd.notna(customer) and customer != '- None -' and str(customer).strip() != '':
            customer_cases = df_original[df_original['Extracted_Customer'] == customer]
            
            # Determine tier based on case count
            if count >= 5:
                tier = "Tier 1"
            elif count >= 3:
                tier = "Tier 2"
            else:
                tier = "Tier 3"
            
            # Priority breakdown
            p1_count = len(customer_cases[customer_cases['Priority'] == 'P1'])
            p2_count = len(customer_cases[customer_cases['Priority'] == 'P2'])
            p3_count = len(customer_cases[customer_cases['Priority'] == 'P3'])
            p4_count = len(customer_cases[customer_cases['Priority'] == 'P4'])
            
            # Status breakdown
            open_cases = len(customer_cases[customer_cases['Status'].isin(['Open', 'In Progress', 'Reopened'])])
            closed_cases = len(customer_cases[customer_cases['Status'].isin(['Closed', 'Done', 'Resolved'])])
            
            # Top resolution and integration
            top_resolution = customer_cases['Resolution'].value_counts().index[0] if len(customer_cases['Resolution'].value_counts()) > 0 else 'N/A'
            top_integration = customer_cases['Custom field (Integration Apps)'].value_counts().index[0] if len(customer_cases['Custom field (Integration Apps)'].value_counts()) > 0 else 'N/A'
            
            # Sample case examples (first 2 case keys)
            sample_cases = customer_cases['Issue key'].head(2).tolist()
            case_examples = ', '.join(sample_cases)
            
            customer_analysis_data.append([
                customer[:25] + '...' if len(customer) > 25 else customer,
                tier,
                str(count),
                str(p1_count),
                str(p2_count),
                str(p3_count),
                str(p4_count),
                str(open_cases),
                str(closed_cases),
                top_resolution[:20] + '...' if len(str(top_resolution)) > 20 else str(top_resolution),
                top_integration[:20] + '...' if len(str(top_integration)) > 20 else str(top_integration),
                case_examples
            ])
    
    customer_analysis_data.append(['', ''])  # Empty row
    
    # Integration analysis section
    customer_analysis_data.append(['INTEGRATION APP ANALYSIS', ''])
    customer_analysis_data.append(['Integration App', 'Total Cases', 'Unique Customers', 'Avg Cases/Customer', 'Top Customer', 'P1 Cases', 'Resolution Rate'])
    
    # Analyze by integration app
    integration_analysis = df_original.groupby('Custom field (Integration Apps)').agg({
        'Issue key': 'count',
        'Extracted_Customer': 'nunique',
        'Priority': lambda x: (x == 'P1').sum(),
        'Resolution': lambda x: (x.isin(['Done', 'Resolved', 'Closed'])).sum()
    }).round(2)
    
    integration_analysis.columns = ['Total_Cases', 'Unique_Customers', 'P1_Cases', 'Resolved_Cases']
    integration_analysis['Avg_Cases_Per_Customer'] = integration_analysis['Total_Cases'] / integration_analysis['Unique_Customers']
    integration_analysis['Resolution_Rate'] = (integration_analysis['Resolved_Cases'] / integration_analysis['Total_Cases'] * 100).round(1)
    
    # Get top customer for each integration
    for integration in integration_analysis.index:
        if pd.notna(integration):
            integration_cases = df_original[df_original['Custom field (Integration Apps)'] == integration]
            # Filter out Unknown customers
            valid_customers = integration_cases[integration_cases['Extracted_Customer'] != 'Unknown']['Extracted_Customer'].value_counts()
            top_customer = valid_customers.index[0] if len(valid_customers) > 0 else 'N/A'
            
            row = integration_analysis.loc[integration]
            customer_analysis_data.append([
                str(integration)[:30] + '...' if len(str(integration)) > 30 else str(integration),
                str(int(row['Total_Cases'])),
                str(int(row['Unique_Customers'])),
                f"{row['Avg_Cases_Per_Customer']:.1f}",
                str(top_customer)[:20] + '...' if len(str(top_customer)) > 20 else str(top_customer),
                str(int(row['P1_Cases'])),
                f"{row['Resolution_Rate']:.1f}%"
            ])
    
    customer_analysis_data.append(['', ''])  # Empty row
    
    # Resolution analysis section
    customer_analysis_data.append(['RESOLUTION ANALYSIS', ''])
    customer_analysis_data.append(['Resolution Type', 'Count', 'Percentage', 'Avg Cases/Customer', 'Top Customer', 'Top Integration'])
    
    resolution_analysis = df_original['Resolution'].value_counts()
    for resolution, count in resolution_analysis.head(15).items():
        if pd.notna(resolution):
            resolution_cases = df_original[df_original['Resolution'] == resolution]
            percentage = (count / total_cases) * 100
            avg_cases = count / len(resolution_cases['Extracted_Customer'].unique())
            # Filter out Unknown customers
            valid_customers = resolution_cases[resolution_cases['Extracted_Customer'] != 'Unknown']['Extracted_Customer'].value_counts()
            top_customer = valid_customers.index[0] if len(valid_customers) > 0 else 'N/A'
            top_integration = resolution_cases['Custom field (Integration Apps)'].value_counts().index[0] if len(resolution_cases['Custom field (Integration Apps)'].value_counts()) > 0 else 'N/A'
            
            customer_analysis_data.append([
                str(resolution)[:25] + '...' if len(str(resolution)) > 25 else str(resolution),
                str(count),
                f"{percentage:.1f}%",
                f"{avg_cases:.1f}",
                str(top_customer)[:20] + '...' if len(str(top_customer)) > 20 else str(top_customer),
                str(top_integration)[:20] + '...' if len(str(top_integration)) > 20 else str(top_integration)
            ])
    
    customer_analysis_df = pd.DataFrame(customer_analysis_data)
    
    print(f"  ✅ Customer extraction complete: {cases_with_customer}/{total_cases} cases ({cases_with_customer/total_cases*100:.1f}%)")
    
    print(f"\n📊 Step 4/4: Combining into Master Report...")
    
    # Read both Excel files
    # For comprehensive sheets with title rows, skip the first row
    comprehensive_sheets = {}
    sheets_with_titles = ['Overall Summary', 'Pattern Analysis', 'Top Integrations', 'All Cases', 'Code Fix with Links']
    
    xls = pd.ExcelFile(temp_comprehensive)
    for sheet_name in xls.sheet_names:
        if sheet_name in sheets_with_titles:
            # Skip the title row (row 0) and use row 1 as headers
            comprehensive_sheets[sheet_name] = pd.read_excel(temp_comprehensive, sheet_name=sheet_name, header=1)
        else:
            comprehensive_sheets[sheet_name] = pd.read_excel(temp_comprehensive, sheet_name=sheet_name)
    
    deep_dive_sheets = pd.read_excel(temp_deep_dive, sheet_name=None)
    
    # Create the master workbook with organized sheet structure
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # SECTION 1: EXECUTIVE SUMMARY (from Comprehensive)
        print(f"  Adding Executive Summary sheets...")
        if 'Overall Summary' in comprehensive_sheets:
            comprehensive_sheets['Overall Summary'].to_excel(writer, sheet_name='Overall Summary', index=False)
        
        # SECTION 2: INTEGRATION ANALYSIS (from Deep Dive)
        print(f"  Adding Integration Analysis sheets...")
        if 'Integration Overview' in deep_dive_sheets:
            deep_dive_sheets['Integration Overview'].to_excel(writer, sheet_name='Integration Overview', index=False)
        if 'Count by Integration' in deep_dive_sheets:
            deep_dive_sheets['Count by Integration'].to_excel(writer, sheet_name='Count by Integration', index=False)
        if 'Resolution Breakdown by IA' in deep_dive_sheets:
            deep_dive_sheets['Resolution Breakdown by IA'].to_excel(writer, sheet_name='Resolution by IA', index=False)
        
        # SECTION 3: TEMPORAL ANALYSIS (from Deep Dive)
        print(f"  Adding Temporal Analysis sheets...")
        if 'Count by Month' in deep_dive_sheets:
            deep_dive_sheets['Count by Month'].to_excel(writer, sheet_name='Count by Month', index=False)
        
        # SECTION 4: CUSTOMER ANALYSIS (Enhanced with Extraction)
        print(f"  Adding Enhanced Customer Analysis sheets...")
        # Use our enhanced customer analysis with extracted data
        customer_analysis_df.to_excel(writer, sheet_name='Customer Analysis Enhanced', index=False, header=False)
        
        # Enhance the original customer analysis with extracted customer details
        print(f"  Enhancing Customer Analysis Original with extracted details...")
        
        # Create a 'Final_Customer_Name' column in df_original
        # This column will prioritize Extracted_Customer if 'Custom field (Customer (old))' is '- None -'
        df_original['Final_Customer_Name'] = df_original.apply(
            lambda row: row['Extracted_Customer'] if row['Custom field (Customer (old))'] == '- None -' else row['Custom field (Customer (old))'],
            axis=1
        )
        
        # Create customer analysis in the format similar to the image
        customer_analysis_data = []
        
        # Get unique customers with their details
        for customer in df_original['Final_Customer_Name'].unique():
            if customer != 'Unknown' and pd.notna(customer):
                customer_cases = df_original[df_original['Final_Customer_Name'] == customer]
                
                # Count cases by priority
                p1_count = len(customer_cases[customer_cases['Priority'] == 'P1'])
                p2_count = len(customer_cases[customer_cases['Priority'] == 'P2'])
                p3_count = len(customer_cases[customer_cases['Priority'] == 'P3'])
                p4_count = len(customer_cases[customer_cases['Priority'] == 'P4'])
                total_cases = len(customer_cases)
                
                # Get top integration app
                top_integration = customer_cases['Custom field (Integration Apps)'].value_counts().index[0] if len(customer_cases['Custom field (Integration Apps)'].value_counts()) > 0 else 'N/A'
                
                # Get most common resolution
                top_resolution = customer_cases['Resolution'].value_counts().index[0] if len(customer_cases['Resolution'].value_counts()) > 0 else 'N/A'
                
                # Determine tier based on case count (similar to image format)
                if total_cases >= 5:
                    tier = "Tier 1"
                elif total_cases >= 3:
                    tier = "Tier 2"
                else:
                    tier = "Tier 3"
                
                customer_analysis_data.append({
                    'Company/Partner Name': customer,
                    'Tier': tier,
                    'Total Cases': total_cases,
                    'P1 Cases': p1_count,
                    'P2 Cases': p2_count,
                    'P3 Cases': p3_count,
                    'P4 Cases': p4_count,
                    'Resolution/Issue Description': top_resolution,
                    'Integration App': top_integration
                })
        
        # Convert to DataFrame and sort by total cases descending
        enhanced_original_customer_analysis_df = pd.DataFrame(customer_analysis_data)
        enhanced_original_customer_analysis_df = enhanced_original_customer_analysis_df.sort_values('Total Cases', ascending=False)
        
        # Write this enhanced version to 'Customer Analysis Original'
        enhanced_original_customer_analysis_df.to_excel(writer, sheet_name='Customer Analysis Original', index=False)
        
        # SECTION 5: TECHNICAL DETAILS (from Deep Dive)
        print(f"  Adding Technical Details sheets...")
        if 'Frequent Flow Issues' in deep_dive_sheets:
            deep_dive_sheets['Frequent Flow Issues'].to_excel(writer, sheet_name='Frequent Flow Issues', index=False)
        if 'Error Categories' in deep_dive_sheets:
            deep_dive_sheets['Error Categories'].to_excel(writer, sheet_name='Error Categories', index=False)
        if 'Error Distribution by IA' in deep_dive_sheets:
            deep_dive_sheets['Error Distribution by IA'].to_excel(writer, sheet_name='Error Distribution', index=False)
        if 'Recurring Errors' in deep_dive_sheets:
            deep_dive_sheets['Recurring Errors'].to_excel(writer, sheet_name='Recurring Errors', index=False)
        
        # SECTION 6: PATTERN ANALYSIS (from Comprehensive)
        print(f"  Adding Pattern Analysis sheets...")
        if 'Pattern Analysis' in comprehensive_sheets:
            comprehensive_sheets['Pattern Analysis'].to_excel(writer, sheet_name='Pattern Analysis', index=False)
        if 'Top Integrations' in comprehensive_sheets:
            comprehensive_sheets['Top Integrations'].to_excel(writer, sheet_name='Top Integrations', index=False)
        
        # SECTION 7: CODE FIX ANALYSIS (from Comprehensive)
        print(f"  Adding Code Fix Analysis sheets...")
        if 'Code Fix with Links' in comprehensive_sheets:
            comprehensive_sheets['Code Fix with Links'].to_excel(writer, sheet_name='Code Fix with Links', index=False)
        
        # SECTION 8: COMPLETE DATA (from both)
        print(f"  Adding Complete Data sheets...")
        if 'All Cases' in comprehensive_sheets:
            comprehensive_sheets['All Cases'].to_excel(writer, sheet_name='All Cases Summary', index=False)
        if 'Case Details' in deep_dive_sheets:
            deep_dive_sheets['Case Details'].to_excel(writer, sheet_name='Complete Case Details', index=False)
    
    # Apply formatting from both source files
    print(f"  Applying formatting...")
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Load the master workbook
    book = load_workbook(output_file)
    
    # Enhanced visual formatting for grand appearance
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')  # Dark blue
    header_font = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
    
    # Alternating row colors
    alt_row_light = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Light gray
    alt_row_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    thick_border = Border(
        left=Side(style='medium', color='1F4E78'),
        right=Side(style='medium', color='1F4E78'),
        top=Side(style='medium', color='1F4E78'),
        bottom=Side(style='medium', color='1F4E78')
    )
    
    # Cell alignment
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for sheet_name in book.sheetnames:
        ws = book[sheet_name]
        
        # Format header row with enhanced style
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thick_border
        
        # Set header row height
        ws.row_dimensions[1].height = 35
        
        # Apply alternating row colors and borders to data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            # Alternating row color
            fill = alt_row_light if row_idx % 2 == 0 else alt_row_white
            
            for col_idx, cell in enumerate(row, start=1):
                cell.border = thin_border
                cell.fill = fill
                
                # Center align numeric columns, left align text
                if isinstance(cell.value, (int, float)):
                    cell.alignment = center_alignment
                    # Format numbers with thousand separators
                    if isinstance(cell.value, int) and cell.value > 999:
                        cell.number_format = '#,##0'
                    elif isinstance(cell.value, float):
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = left_alignment
                
                # Set font for data cells
                cell.font = Font(name='Calibri', size=11)
        
        # Auto-adjust column widths with better sizing
        for idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column = get_column_letter(idx)
            for cell in col:
                try:
                    if cell.value:
                        # Add extra space for better readability
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            # Set width with min and max constraints
            adjusted_width = min(max(max_length + 4, 12), 80)
            ws.column_dimensions[column].width = adjusted_width
        
        # Freeze panes (freeze first row and first column)
        ws.freeze_panes = 'B2'
        
        # Add filter to header row
        ws.auto_filter.ref = ws.dimensions
    
    book.save(output_file)
    
    # Clean up temporary files
    import os
    try:
        os.remove(temp_comprehensive)
        os.remove(temp_deep_dive)
    except:
        pass
    
    print(f"\n{'='*100}")
    print(f"✅ MASTER REPORT GENERATED SUCCESSFULLY!")
    print(f"{'='*100}")
    print(f"\n📁 Output file: {output_file}")
    print(f"\n📋 Master report contains 16 beautifully formatted sheets:")
    print(f"\n  📊 EXECUTIVE SUMMARY")
    print(f"    • Overall Summary")
    print(f"\n  🔧 INTEGRATION ANALYSIS")
    print(f"    • Integration Overview")
    print(f"    • Count by Integration")
    print(f"    • Resolution by IA")
    print(f"\n  📅 TEMPORAL ANALYSIS")
    print(f"    • Count by Month")
    print(f"\n  👥 CUSTOMER ANALYSIS (ENHANCED)")
    print(f"    • Customer Analysis Enhanced (with extracted customer info)")
    print(f"    • Customer Analysis Original (from deep dive)")
    print(f"\n  🔍 TECHNICAL DETAILS")
    print(f"    • Frequent Flow Issues")
    print(f"    • Error Categories")
    print(f"    • Error Distribution")
    print(f"    • Recurring Errors")
    print(f"\n  📈 PATTERN ANALYSIS")
    print(f"    • Pattern Analysis")
    print(f"    • Top Integrations")
    print(f"\n  ✅ CODE FIX ANALYSIS")
    print(f"    • Code Fix with Links")
    print(f"\n  📝 COMPLETE DATA")
    print(f"    • All Cases Summary")
    print(f"    • Complete Case Details")
    print(f"\n  ✨ VISUAL ENHANCEMENTS:")
    print(f"    ✓ Dark blue headers with white text (size 12)")
    print(f"    ✓ Alternating row colors (gray/white)")
    print(f"    ✓ Professional borders on all cells")
    print(f"    ✓ Auto-filters on all sheets")
    print(f"    ✓ Freeze panes (row 1, column A)")
    print(f"    ✓ Auto-sized columns with min/max constraints")
    print(f"    ✓ Number formatting with thousand separators")
    print(f"    ✓ Center-aligned numbers, left-aligned text")
    print(f"\n{'='*100}")
    print(f"🎯 ONE REPORT TO RULE THEM ALL!")
    print(f"{'='*100}")
    print(f"\n📊 CUSTOMER EXTRACTION SUMMARY:")
    print(f"   • Total cases analyzed: {total_cases}")
    print(f"   • Customer info extracted: {cases_with_customer} cases ({cases_with_customer/total_cases*100:.1f}%)")
    print(f"   • Cases without customer info: {cases_without_customer} cases ({cases_without_customer/total_cases*100:.1f}%)")
    print(f"   • Unique customers identified: {len(customer_stats)}")
    print(f"\n🔍 KEY IMPROVEMENT:")
    print(f"   • Before: {cases_without_customer/total_cases*100:.1f}% cases appeared to have 'no customer details'")
    print(f"   • After: {cases_with_customer/total_cases*100:.1f}% cases now have customer information!")
    print(f"   • Customer names extracted from 'Company:' field in descriptions")
    
    return True

# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description='Master Report Generator - Single comprehensive file with ALL insights',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  python3 generate_master_report.py --file 2025.csv
  python3 generate_master_report.py --file /path/to/2024.csv --output 2024_Complete_Report.xlsx
  
This tool generates ONE master Excel file with 15 organized sheets:
  • Executive Summary (Overall Summary)
  • Integration Analysis (Overview, Counts, Resolutions)
  • Temporal Analysis (Monthly trends)
  • Customer Analysis (Customer-specific data)
  • Technical Details (Flows, Errors, Recurring issues)
  • Pattern Analysis (Patterns, Top integrations)
  • Code Fix Analysis (Done cases with linked PRE/PRD tickets)
  • Complete Data (All cases + detailed extractions)
        '''
    )
    
    parser.add_argument('--file', '-f', required=True, help='Path to CSV file with cases')
    parser.add_argument('--output', '-o', help='Output Excel filename (optional, auto-generated if not provided)')
    parser.add_argument('--version', action='version', version=f'%(prog)s {VERSION}')
    
    args = parser.parse_args()
    
    try:
        success = generate_master_report(args.file, args.output)
        if not success:
            sys.exit(1)
    except FileNotFoundError:
        print(f"\n❌ Error: File not found: {args.file}")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Error during analysis: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

