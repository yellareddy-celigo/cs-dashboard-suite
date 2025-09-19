#!/usr/bin/env python3
"""
Advanced Clear Dashboard Creator
Creates a professional dashboard with charts and visual elements similar to your Google Sheet
"""

import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.utils import get_column_letter

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class AdvancedDashboardCreator:
    def __init__(self):
        """Initialize the dashboard creator"""
        self.cs_data = None
        self.pre_data = None
        
    def connect_to_google_sheets(self):
        """Connect to your Google Sheet"""
        try:
            scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
            credentials = Credentials.from_service_account_file('service_account_key.json', scopes=scopes)
            self.gc = gspread.authorize(credentials)
            
            sheet_id = '1JN9HWj8JvClgTDe77x3KIxuyNUTsybPhYhhfzKy6Udc'
            self.sheet = self.gc.open_by_key(sheet_id)
            logger.info("✅ Connected to Google Sheets successfully")
            return True
        except Exception as e:
            logger.error(f"❌ Google Sheets connection failed: {e}")
            return False
    
    def load_data(self):
        """Load data from your Google Sheet"""
        try:
            # Load CS data
            cs_sheet = self.sheet.worksheet('Raw Data - CS')
            cs_values = cs_sheet.get_all_values()
            cs_headers = cs_values[0]
            cs_data = cs_values[1:]
            
            # Create DataFrame
            self.cs_data = pd.DataFrame(cs_data, columns=cs_headers)
            logger.info(f"✅ Loaded {len(self.cs_data)} CS issues")
            
            # Load PRE data
            pre_sheet = self.sheet.worksheet('Raw Data - PRE')
            pre_values = pre_sheet.get_all_values()
            pre_headers = pre_values[0]
            pre_data = pre_values[1:]
            
            # Create DataFrame
            self.pre_data = pd.DataFrame(pre_data, columns=pre_headers)
            logger.info(f"✅ Loaded {len(self.pre_data)} PRE issues")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Error loading data: {e}")
            return False
    
    def process_dates(self, df):
        """Process date columns"""
        date_columns = ['Created', 'Updated', 'Resolved']
        
        for col in date_columns:
            if col in df.columns:
                try:
                    df[col + '_dt'] = pd.to_datetime(df[col], errors='coerce')
                    df[col + '_year'] = df[col + '_dt'].dt.year
                    df[col + '_month'] = df[col + '_dt'].dt.month
                    df[col + '_month_name'] = df[col + '_dt'].dt.strftime('%B')
                except:
                    pass
        
        return df
    
    def generate_analysis(self):
        """Generate comprehensive analysis"""
        analysis = {}
        
        # Process dates
        self.cs_data = self.process_dates(self.cs_data)
        self.pre_data = self.process_dates(self.pre_data)
        
        # Basic statistics
        analysis['total_issues'] = len(self.cs_data) + len(self.pre_data)
        analysis['cs_issues'] = len(self.cs_data)
        analysis['pre_issues'] = len(self.pre_data)
        
        # Status analysis
        cs_status = self.cs_data['Status'].value_counts().to_dict() if 'Status' in self.cs_data.columns else {}
        pre_status = self.pre_data['Status'].value_counts().to_dict() if 'Status' in self.pre_data.columns else {}
        
        # Combine status counts
        all_status = {}
        for status, count in cs_status.items():
            all_status[status] = all_status.get(status, 0) + count
        for status, count in pre_status.items():
            all_status[status] = all_status.get(status, 0) + count
        
        analysis['status_distribution'] = all_status
        analysis['cs_status_distribution'] = cs_status
        analysis['pre_status_distribution'] = pre_status
        
        # Monthly trends
        cs_monthly = self.cs_data['Created_month_name'].value_counts().to_dict() if 'Created_month_name' in self.cs_data.columns else {}
        pre_monthly = self.pre_data['Created_month_name'].value_counts().to_dict() if 'Created_month_name' in self.pre_data.columns else {}
        
        # Combine monthly counts
        all_monthly = {}
        for month, count in cs_monthly.items():
            all_monthly[month] = all_monthly.get(month, 0) + count
        for month, count in pre_monthly.items():
            all_monthly[month] = all_monthly.get(month, 0) + count
        
        analysis['monthly_trends'] = all_monthly
        analysis['cs_monthly_trends'] = cs_monthly
        analysis['pre_monthly_trends'] = pre_monthly
        
        # Priority analysis
        if 'Priority' in self.cs_data.columns:
            cs_priority = self.cs_data['Priority'].value_counts().to_dict()
            analysis['cs_priority_distribution'] = cs_priority
        
        if 'Priority' in self.pre_data.columns:
            pre_priority = self.pre_data['Priority'].value_counts().to_dict()
            analysis['pre_priority_distribution'] = pre_priority
        
        # Assignee analysis
        if 'Assignee' in self.cs_data.columns:
            cs_assignee = self.cs_data['Assignee'].value_counts().head(10).to_dict()
            analysis['cs_top_assignees'] = cs_assignee
        
        if 'Assignee' in self.pre_data.columns:
            pre_assignee = self.pre_data['Assignee'].value_counts().head(10).to_dict()
            analysis['pre_top_assignees'] = pre_assignee
        
        logger.info("✅ Analysis completed")
        return analysis
    
    def create_ascii_chart(self, value, max_value, width=20):
        """Create ASCII chart bar"""
        if max_value == 0:
            return '⬜' * width
        
        bar_length = int((value / max_value) * width)
        return '🟨' * bar_length + '⬜' * (width - bar_length)
    
    def create_excel_dashboard(self, analysis, filename=None):
        """Create an advanced Excel dashboard with charts"""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'advanced_dashboard_{timestamp}.xlsx'
        
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Define styles
            title_font = Font(bold=True, size=18, color="FFFFFF")
            header_font = Font(bold=True, size=12, color="FFFFFF")
            subheader_font = Font(bold=True, size=11, color="FFFFFF")
            data_font = Font(size=10)
            
            title_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
            subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cs_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
            pre_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # 1. MAIN VISUAL OVERVIEW (like your Google Sheet)
            ws_main = wb.create_sheet("🎨 Main Visual Overview")
            
            # Create the main overview similar to your Google Sheet
            main_data = [
                ['🚨 MAIN VISUAL OVERVIEW - HISTORICAL CS ANALYSIS 🚨', '', '', ''],
                ['═════════════════════════════════════════════════════════════════════════', '', '', ''],
                [f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | Total Issues: {analysis["total_issues"]} | CS: {analysis["cs_issues"]} | PRE: {analysis["pre_issues"]}', '', '', ''],
                ['═════════════════════════════════════════════════════════════════════════', '', '', ''],
                ['', '', '', ''],
                ['📊 BIG PICTURE: MONTHLY ISSUE DISTRIBUTION', '', '', ''],
                ['', '', '', ''],
                ['Month', 'Total', 'CS', 'PRE', 'Visual Chart']
            ]
            
            months = ['January', 'February', 'March', 'April', 'May', 'June', 
                     'July', 'August', 'September', 'October', 'November', 'December']
            
            max_monthly = max(analysis['monthly_trends'].values()) if analysis['monthly_trends'] else 1
            
            for month in months:
                total = analysis['monthly_trends'].get(month, 0)
                cs = analysis['cs_monthly_trends'].get(month, 0)
                pre = analysis['pre_monthly_trends'].get(month, 0)
                chart = self.create_ascii_chart(total, max_monthly)
                main_data.append([month, total, cs, pre, chart])
            
            # Add status breakdown
            main_data.extend([
                ['', '', '', '', ''],
                ['📋 STATUS BREAKDOWN', '', '', '', ''],
                ['Status', 'Total', 'CS', 'PRE', 'Percentage']
            ])
            
            for status, count in analysis['status_distribution'].items():
                cs_count = analysis['cs_status_distribution'].get(status, 0)
                pre_count = analysis['pre_status_distribution'].get(status, 0)
                percentage = (count / analysis['total_issues']) * 100
                main_data.append([status, count, cs_count, pre_count, f"{percentage:.1f}%"])
            
            # Write main data
            for row in main_data:
                ws_main.append(row)
            
            # Style the main sheet
            for cell in ws_main[1]:
                cell.font = title_font
                cell.fill = title_fill
                cell.border = border
            
            for cell in ws_main[2]:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # 2. EXECUTIVE DASHBOARD
            ws_exec = wb.create_sheet("🎯 Executive Dashboard")
            
            exec_data = [
                ['🎯 EXECUTIVE DASHBOARD - KEY PERFORMANCE INDICATORS', '', '', ''],
                ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '', ''],
                ['', '', '', ''],
                ['📊 KEY METRICS OVERVIEW', '', '', ''],
                ['Metric', 'Value', 'CS', 'PRE'],
                ['Total Issues', analysis['total_issues'], analysis['cs_issues'], analysis['pre_issues']],
                ['CS Percentage', f"{(analysis['cs_issues']/analysis['total_issues']*100):.1f}%", '', ''],
                ['PRE Percentage', f"{(analysis['pre_issues']/analysis['total_issues']*100):.1f}%", '', ''],
                ['', '', '', ''],
                ['📈 MONTHLY TRENDS WITH VISUAL CHARTS', '', '', ''],
                ['Month', 'Total', 'CS', 'PRE', 'Visual Chart']
            ]
            
            for month in months:
                total = analysis['monthly_trends'].get(month, 0)
                cs = analysis['cs_monthly_trends'].get(month, 0)
                pre = analysis['pre_monthly_trends'].get(month, 0)
                chart = self.create_ascii_chart(total, max_monthly)
                exec_data.append([month, total, cs, pre, chart])
            
            for row in exec_data:
                ws_exec.append(row)
            
            # Style executive dashboard
            for cell in ws_exec[1]:
                cell.font = title_font
                cell.fill = title_fill
                cell.border = border
            
            # 3. DETAILED ANALYSIS SHEET
            ws_detail = wb.create_sheet("📊 Detailed Analysis")
            
            detail_data = [
                ['📊 DETAILED ISSUE ANALYSIS - COMPREHENSIVE BREAKDOWN', '', '', ''],
                ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '', ''],
                ['', '', '', ''],
                ['🔍 CS vs PRE COMPARISON', '', '', ''],
                ['Metric', 'CS', 'PRE', 'Difference'],
                ['Total Issues', analysis['cs_issues'], analysis['pre_issues'], 
                 analysis['cs_issues'] - analysis['pre_issues']],
                ['', '', '', ''],
                ['📋 TOP ASSIGNEES', '', '', ''],
                ['Assignee', 'Total Issues', 'CS Issues', 'PRE Issues']
            ]
            
            # Combine assignee data
            all_assignees = {}
            if 'cs_top_assignees' in analysis:
                for assignee, count in analysis['cs_top_assignees'].items():
                    all_assignees[assignee] = all_assignees.get(assignee, 0) + count
            if 'pre_top_assignees' in analysis:
                for assignee, count in analysis['pre_top_assignees'].items():
                    all_assignees[assignee] = all_assignees.get(assignee, 0) + count
            
            for assignee, count in sorted(all_assignees.items(), key=lambda x: x[1], reverse=True)[:10]:
                cs_count = analysis['cs_top_assignees'].get(assignee, 0)
                pre_count = analysis['pre_top_assignees'].get(assignee, 0)
                detail_data.append([assignee, count, cs_count, pre_count])
            
            for row in detail_data:
                ws_detail.append(row)
            
            # Style detailed analysis
            for cell in ws_detail[1]:
                cell.font = title_font
                cell.fill = title_fill
                cell.border = border
            
            # 4. RAW DATA SHEET
            ws_raw = wb.create_sheet("📋 Raw Data")
            
            # Add CS data
            cs_data_with_type = self.cs_data.copy()
            cs_data_with_type['Project Type'] = 'CS'
            
            # Select key columns for display
            key_columns = ['Project Type', 'Summary', 'Issue key', 'Status', 'Priority', 'Created', 'Updated', 'Assignee']
            available_columns = [col for col in key_columns if col in cs_data_with_type.columns]
            
            for r in dataframe_to_rows(cs_data_with_type[available_columns], index=False, header=True):
                ws_raw.append(r)
            
            # Add PRE data
            pre_data_with_type = self.pre_data.copy()
            pre_data_with_type['Project Type'] = 'PRE'
            
            available_pre_columns = [col for col in key_columns if col in pre_data_with_type.columns]
            
            for r in dataframe_to_rows(pre_data_with_type[available_pre_columns], index=False, header=False):
                ws_raw.append(r)
            
            # Style raw data
            for cell in ws_raw[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Color code by project type
            for row_idx, row in enumerate(ws_raw.iter_rows(min_row=2), 2):
                project_type = ws_raw.cell(row=row_idx, column=1).value
                for cell in row:
                    if project_type == 'CS':
                        cell.fill = cs_fill
                    elif project_type == 'PRE':
                        cell.fill = pre_fill
                    cell.border = border
            
            # 5. CHARTS AND VISUALIZATIONS SHEET
            ws_charts = wb.create_sheet("📈 Charts & Visualizations")
            
            charts_data = [
                ['📈 VISUAL CHARTS & INSIGHTS', '', '', ''],
                ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '', ''],
                ['', '', '', ''],
                ['📊 MONTHLY ISSUE DISTRIBUTION WITH ASCII CHARTS', '', '', ''],
                ['Month', 'Total Issues', 'CS Issues', 'PRE Issues', 'Visual Chart']
            ]
            
            for month in months:
                total = analysis['monthly_trends'].get(month, 0)
                cs = analysis['cs_monthly_trends'].get(month, 0)
                pre = analysis['pre_monthly_trends'].get(month, 0)
                chart = self.create_ascii_chart(total, max_monthly)
                charts_data.append([month, total, cs, pre, chart])
            
            charts_data.extend([
                ['', '', '', '', ''],
                ['📋 STATUS DISTRIBUTION', '', '', '', ''],
                ['Status', 'Count', 'Percentage', 'CS', 'PRE']
            ])
            
            for status, count in analysis['status_distribution'].items():
                percentage = (count / analysis['total_issues']) * 100
                cs_count = analysis['cs_status_distribution'].get(status, 0)
                pre_count = analysis['pre_status_distribution'].get(status, 0)
                charts_data.append([status, count, f"{percentage:.1f}%", cs_count, pre_count])
            
            for row in charts_data:
                ws_charts.append(row)
            
            # Style charts sheet
            for cell in ws_charts[1]:
                cell.font = title_font
                cell.fill = title_fill
                cell.border = border
            
            # Auto-adjust column widths
            for ws in [ws_main, ws_exec, ws_detail, ws_raw, ws_charts]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the file
            wb.save(filename)
            logger.info(f"✅ Advanced Excel dashboard saved: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"❌ Error creating Excel dashboard: {e}")
            return None
    
    def run(self):
        """Run the complete dashboard creation process"""
        logger.info("🚀 Starting Advanced Clear Dashboard Creation")
        logger.info("=" * 50)
        
        # Connect to Google Sheets
        if not self.connect_to_google_sheets():
            return False
        
        # Load data
        if not self.load_data():
            return False
        
        # Generate analysis
        analysis = self.generate_analysis()
        
        # Create Excel dashboard
        filename = self.create_excel_dashboard(analysis)
        
        if filename:
            logger.info("✅ Advanced clear dashboard created successfully!")
            logger.info(f"📁 File: {filename}")
            logger.info(f"📊 Analyzed {analysis['total_issues']} total issues")
            logger.info(f"   • CS Issues: {analysis['cs_issues']}")
            logger.info(f"   • PRE Issues: {analysis['pre_issues']}")
            return True
        else:
            logger.error("❌ Dashboard creation failed")
            return False

def main():
    """Main execution function"""
    creator = AdvancedDashboardCreator()
    success = creator.run()
    
    if success:
        print("\n🎉 SUCCESS! Advanced clear dashboard created!")
        print("📋 Dashboard includes:")
        print("   • 🎨 Main Visual Overview - Similar to your Google Sheet")
        print("   • 🎯 Executive Dashboard - Key metrics with visual charts")
        print("   • 📊 Detailed Analysis - Comprehensive breakdown")
        print("   • 📋 Raw Data - All issues with color coding")
        print("   • 📈 Charts & Visualizations - ASCII charts and insights")
        print("\n🔗 Open the Excel file to view your clear dashboard!")
    else:
        print("\n❌ Dashboard creation failed. Check the logs above for details.")
    
    return 0 if success else 1

if __name__ == "__main__":
    import sys
    sys.exit(main())
